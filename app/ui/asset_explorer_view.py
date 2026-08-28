import sqlite3
from pathlib import Path
from datetime import datetime

from PySide6.QtCore import Qt
from PySide6.QtWidgets import (
    QWidget,
    QVBoxLayout,
    QHBoxLayout,
    QLabel,
    QTreeWidget,
    QTreeWidgetItem,
    QFrame,
    QLineEdit,
    QComboBox,
    QPushButton,
    QSplitter,
    QScrollArea,
    QTableWidget,
    QTableWidgetItem,
    QHeaderView,
    QFileDialog,
    QMessageBox,
)

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

from app.config.settings import PROJECTS_DIR


class AssetExplorerView(QWidget):

    """
    GLOBAL ASSET MANAGEMENT

    Physical asset identity:
        asset_id

    Project occurrence identity:
        node_id

    Project-local hierarchy:
        parent_id

    Global hierarchy:
        asset.metadata["parent_asset_id"]

    This distinction is important when an asset is linked into
    another project.

    Example:

        Project 1

        REF-3 SS-1
        ├── HV-201A
        └── HV-201B


        Project 2

        REF-3 SS-1
        └── HV-201C


    Global Asset Management:

        REF-3 SS-1
        ├── HV-201A
        ├── HV-201B
        └── HV-201C
    """

    def __init__(
        self,
        global_asset_service,
        parent=None
    ):

        super().__init__(parent)

        self.global_asset_service = (
            global_asset_service
        )

        self._all_records = []

        self._build_ui()

        self.refresh()

    # =========================================================
    # UI
    # =========================================================

    def _build_ui(self):

        root = QVBoxLayout(self)

        root.setContentsMargins(
            16,
            16,
            16,
            16
        )

        root.setSpacing(10)

        # -----------------------------------------------------
        # HEADER
        # -----------------------------------------------------

        header_layout = QHBoxLayout()

        header = QLabel(
            "Asset Management"
        )

        header.setStyleSheet(
            """
            QLabel {
                font-size: 24px;
                font-weight: 700;
            }
            """
        )

        header_layout.addWidget(header)

        header_layout.addStretch()

        self.export_button = QPushButton(
            "Export Asset Register"
        )

        self.export_button.clicked.connect(
            self.export_asset_register
        )

        header_layout.addWidget(
            self.export_button
        )

        root.addLayout(
            header_layout
        )

        subtitle = QLabel(
            "Global physical asset configuration, hierarchy and test history"
        )

        subtitle.setStyleSheet(
            """
            QLabel {
                color: #999999;
                font-size: 13px;
            }
            """
        )

        root.addWidget(
            subtitle
        )

        # -----------------------------------------------------
        # FILTER BAR
        # -----------------------------------------------------

        filter_frame = QFrame()

        filter_layout = QHBoxLayout(
            filter_frame
        )

        filter_layout.setContentsMargins(
            8,
            6,
            8,
            6
        )

        self.search_edit = QLineEdit()

        self.search_edit.setPlaceholderText(
            "Search asset, tag, manufacturer, model, serial number..."
        )

        self.search_edit.textChanged.connect(
            self.apply_filters
        )

        filter_layout.addWidget(
            self.search_edit,
            1
        )

        self.type_filter = QComboBox()

        self.type_filter.addItems(
            [
                "All Assets",
                "Substations",
                "Switchboards",
                "Panels",
                "Components",
            ]
        )

        self.type_filter.currentIndexChanged.connect(
            self.apply_filters
        )

        filter_layout.addWidget(
            self.type_filter
        )

        expand_button = QPushButton(
            "Expand All"
        )

        expand_button.clicked.connect(
            self.expand_all
        )

        filter_layout.addWidget(
            expand_button
        )

        collapse_button = QPushButton(
            "Collapse All"
        )

        collapse_button.clicked.connect(
            self.collapse_all
        )

        filter_layout.addWidget(
            collapse_button
        )

        refresh_button = QPushButton(
            "Refresh"
        )

        refresh_button.clicked.connect(
            self.refresh
        )

        filter_layout.addWidget(
            refresh_button
        )

        root.addWidget(
            filter_frame
        )

        # -----------------------------------------------------
        # SPLITTER
        # -----------------------------------------------------

        splitter = QSplitter(
            Qt.Orientation.Horizontal
        )

        splitter.setChildrenCollapsible(
            False
        )

        # -----------------------------------------------------
        # TREE
        # -----------------------------------------------------

        tree_frame = QFrame()

        tree_layout = QVBoxLayout(
            tree_frame
        )

        tree_title = QLabel(
            "Asset Structure"
        )

        tree_title.setStyleSheet(
            """
            QLabel {
                font-size: 16px;
                font-weight: 700;
                padding: 6px;
            }
            """
        )

        tree_layout.addWidget(
            tree_title
        )

        self.tree = QTreeWidget()

        self.tree.setHeaderLabels(
            [
                "Asset",
                "Type",
                "Asset Tag",
            ]
        )

        self.tree.setIndentation(
            22
        )

        self.tree.setAnimated(
            True
        )

        self.tree.setUniformRowHeights(
            True
        )

        self.tree.setColumnWidth(
            0,
            360
        )

        self.tree.itemSelectionChanged.connect(
            self._selection_changed
        )

        tree_layout.addWidget(
            self.tree
        )

        splitter.addWidget(
            tree_frame
        )

        # -----------------------------------------------------
        # DETAILS
        # -----------------------------------------------------

        details_frame = QFrame()

        details_layout = QVBoxLayout(
            details_frame
        )

        self.details_title = QLabel(
            "Select an asset"
        )

        self.details_title.setStyleSheet(
            """
            QLabel {
                font-size: 21px;
                font-weight: 700;
                padding: 8px;
            }
            """
        )

        details_layout.addWidget(
            self.details_title
        )

        self.details_subtitle = QLabel(
            ""
        )

        self.details_subtitle.setStyleSheet(
            """
            QLabel {
                color: #999999;
                padding: 0 8px 8px 8px;
            }
            """
        )

        details_layout.addWidget(
            self.details_subtitle
        )

        scroll = QScrollArea()

        scroll.setWidgetResizable(
            True
        )

        scroll.setFrameShape(
            QFrame.Shape.NoFrame
        )

        self.details_container = QWidget()

        self.details_form_layout = QVBoxLayout(
            self.details_container
        )

        self.details_form_layout.setContentsMargins(
            8,
            8,
            8,
            8
        )

        self.details_form_layout.setSpacing(
            8
        )

        scroll.setWidget(
            self.details_container
        )

        details_layout.addWidget(
            scroll
        )

        splitter.addWidget(
            details_frame
        )

        splitter.setStretchFactor(
            0,
            5
        )

        splitter.setStretchFactor(
            1,
            5
        )

        root.addWidget(
            splitter,
            1
        )

        # -----------------------------------------------------
        # STYLE
        # -----------------------------------------------------

        self.setStyleSheet(
            """
            QFrame {
                background: #292929;
                border-radius: 7px;
            }

            QTreeWidget {
                background: #242424;
                border: 1px solid #3d3d3d;
                border-radius: 7px;
                outline: none;
            }

            QTreeWidget::item {
                padding: 6px;
                min-height: 28px;
            }

            QTreeWidget::item:hover {
                background: #333333;
            }

            QTreeWidget::item:selected {
                background: #3b3b3b;
                border-left: 3px solid #f39c12;
            }

            QLineEdit,
            QComboBox {
                min-height: 32px;
                padding: 4px 8px;
                border: 1px solid #444444;
                border-radius: 6px;
                background: #303030;
            }

            QPushButton {
                min-height: 32px;
                padding: 5px 12px;
                border: 1px solid #444444;
                border-radius: 6px;
                background: #303030;
            }

            QPushButton:hover {
                background: #3a3a3a;
            }

            QLabel#Section {
                font-size: 15px;
                font-weight: 700;
                padding: 7px 4px;
            }

            QLabel#Value {
                padding: 7px;
                background: #303030;
                border-radius: 5px;
            }

            QTableWidget {
                background: #242424;
                border: 1px solid #444444;
            }
            """
        )

    # =========================================================
    # REFRESH
    # =========================================================

    def refresh(self):

        try:

            self.global_asset_service.refresh()

            self._load_records()

            self.apply_filters()

        except Exception as error:

            QMessageBox.critical(
                self,
                "Asset Management",
                f"Unable to refresh assets:\n\n{error}"
            )

    # =========================================================
    # LOAD RECORDS
    # =========================================================

    def _load_records(self):

        self._all_records = []

        # =====================================================
        # NODES
        #
        # IMPORTANT:
        #
        # Every project-local node is retained separately.
        #
        # node_id   -> identifies the node in the project tree
        # parent_id -> defines the tree hierarchy
        # asset_id  -> identifies the physical/global asset
        #
        # DO NOT deduplicate nodes using asset_id.
        # =====================================================

        for entry in self.global_asset_service.get_all_nodes():

            node = entry.get("node")

            if node is None:
                continue

            node_type = self._normalise_type(
                getattr(
                    node,
                    "node_type",
                    ""
                )
            )

            if node_type not in (
                "SUBSTATION",
                "SWITCHBOARD",
                "PANEL",
            ):
                continue

            record = {
                "record_type": "NODE",

                # -------------------------------------------------
                # Project information
                # -------------------------------------------------

                "project": entry.get(
                    "project",
                    ""
                ),

                "folder": entry.get(
                    "folder"
                ),

                # -------------------------------------------------
                # Managers
                # -------------------------------------------------

                "asset_manager": entry.get(
                    "asset_manager"
                ),

                "component_manager": entry.get(
                    "component_manager"
                ),

                # -------------------------------------------------
                # Actual node
                # -------------------------------------------------

                "node": node,

                # -------------------------------------------------
                # Tree identity
                # -------------------------------------------------

                "node_id": getattr(
                    node,
                    "node_id",
                    None
                ),

                "parent_id": getattr(
                    node,
                    "parent_id",
                    None
                ),

                # -------------------------------------------------
                # Physical/global identity
                # -------------------------------------------------

                "asset_id": getattr(
                    node,
                    "asset_id",
                    None
                ),

                # -------------------------------------------------
                # Display
                # -------------------------------------------------

                "name": getattr(
                    node,
                    "name",
                    ""
                ),

                "type": node_type,
            }

            self._all_records.append(
                record
            )

        # =====================================================
        # COMPONENTS
        # =====================================================

        for entry in self.global_asset_service.get_all_components():

            component = entry.get(
                "component"
            )

            panel = entry.get(
                "panel"
            )

            if component is None:
                continue

            component_type = (
                self._normalise_component_type(
                    getattr(
                        component,
                        "component_type",
                        ""
                    )
                )
            )

            record = {
                "record_type": "COMPONENT",

                "project": entry.get(
                    "project",
                    ""
                ),

                "folder": entry.get(
                    "folder"
                ),

                "asset_manager": entry.get(
                    "asset_manager"
                ),

                "component_manager": entry.get(
                    "component_manager"
                ),

                "node": panel,

                "component": component,

                "type": component_type,

                # The panel's physical asset.
                "asset_id": getattr(
                    panel,
                    "asset_id",
                    None
                ),

                # The panel's project-local node.
                "panel_node_id": getattr(
                    panel,
                    "node_id",
                    None
                ),

                "component_id": getattr(
                    component,
                    "component_id",
                    None
                ),

                "name": getattr(
                    component,
                    "name",
                    ""
                ),
            }

            self._all_records.append(
                record
            )
    # =========================================================
    # MERGE CONFIGURATION
    # =========================================================

    @staticmethod
    def _merge_node_configuration(
        target,
        source
    ):

        fields = (

            "name",
            "asset_tag",
            "equipment_name",
            "equipment_type",
            "manufacturer",
            "model",
            "serial_number",
            "ct_count",
            "relay_count",
            "aux_count",
            "meter_count",

        )

        for field in fields:

            current = getattr(
                target,
                field,
                None
            )

            incoming = getattr(
                source,
                field,
                None
            )

            if (
                not current
                and incoming
            ):

                try:

                    setattr(
                        target,
                        field,
                        incoming
                    )

                except Exception:

                    pass

    # =========================================================
    # FILTER / BUILD TREE
    # =========================================================

    def apply_filters(self):

        self.tree.clear()
        self._tree_nodes = {}

        search = (
            self.search_edit.text()
            .strip()
            .lower()
        )

        selected_type = (
            self.type_filter.currentText()
        )

        # =====================================================
        # GET ALL NODE RECORDS
        # =====================================================

        all_nodes = [
            record
            for record in self._all_records
            if record.get("record_type") == "NODE"
        ]

        all_components = [
            record
            for record in self._all_records
            if record.get("record_type") == "COMPONENT"
        ]

        # =====================================================
        # STEP 1
        #
        # Build a lookup of every project-local node.
        #
        # IMPORTANT:
        #
        # parent_id is used here.
        #
        # A parent_id only has meaning inside the project
        # in which that node exists.
        #
        # Therefore the lookup key is:
        #
        #       (project, node_id)
        # =====================================================

        local_nodes = {}

        for record in all_nodes:

            project = str(
                record.get(
                    "project",
                    ""
                )
            )

            node_id = record.get(
                "node_id"
            )

            if not node_id:
                continue

            local_nodes[
                (project, node_id)
            ] = record

        # =====================================================
        # STEP 2
        #
        # Resolve every project-local node into a GLOBAL
        # PHYSICAL NODE KEY.
        #
        # This is the important bit.
        #
        # parent_id builds the original hierarchy.
        #
        # asset_id merges copies of the same physical asset.
        #
        # Example:
        #
        # Project A:
        #
        # REF-3 SS-1
        # node_id = AAA
        #
        # Project B:
        #
        # REF-3 SS-1
        # node_id = BBB
        #
        # They can have different node_ids but represent the
        # same physical asset.
        # =====================================================

        global_nodes = {}

        # Maps:
        #
        # (project, node_id)
        #
        #       ->
        #
        # global_key
        #
        local_to_global = {}

        # -----------------------------------------------------
        # First pass: create the global node records.
        # -----------------------------------------------------

        for record in all_nodes:

            global_key = self._get_global_node_key(
                record
            )

            local_key = (
                str(
                    record.get(
                        "project",
                        ""
                    )
                ),
                record.get(
                    "node_id"
                )
            )

            local_to_global[
                local_key
            ] = global_key

            # -------------------------------------------------
            # Create the global node only once.
            # -------------------------------------------------

            if global_key not in global_nodes:

                global_nodes[
                    global_key
                ] = {
                    "record_type": "NODE",

                    "node": record.get(
                        "node"
                    ),

                    "name": record.get(
                        "name",
                        ""
                    ),

                    "type": record.get(
                        "type",
                        ""
                    ),

                    "asset_id": record.get(
                        "asset_id"
                    ),

                    "node_id": record.get(
                        "node_id"
                    ),

                    "parent_id": None,

                    "project": record.get(
                        "project",
                        ""
                    ),

                    "records": [
                        record
                    ],
                }

            else:

                # -------------------------------------------------
                # Same physical asset encountered in another
                # project.
                #
                # Keep all source records attached to it.
                # -------------------------------------------------

                global_nodes[
                    global_key
                ][
                    "records"
                ].append(
                    record
                )

        # =====================================================
        # STEP 3
        #
        # Determine the GLOBAL parent of every global node.
        #
        # We STILL use the original parent_id.
        #
        # But instead of attaching:
        #
        #     child -> raw parent_id
        #
        # we resolve:
        #
        #     project + parent_id
        #
        # into the parent's GLOBAL KEY.
        # =====================================================

        for global_key, global_record in global_nodes.items():

            parent_global_keys = []

            for source_record in global_record[
                "records"
            ]:

                project = str(
                    source_record.get(
                        "project",
                        ""
                    )
                )

                parent_id = source_record.get(
                    "parent_id"
                )

                # -------------------------------------------------
                # Root node.
                # -------------------------------------------------

                if (
                    parent_id is None
                    or parent_id == ""
                ):

                    continue

                parent_local_key = (
                    project,
                    parent_id
                )

                parent_global_key = local_to_global.get(
                    parent_local_key
                )

                if parent_global_key is not None:

                    parent_global_keys.append(
                        parent_global_key
                    )

            # -----------------------------------------------------
            # If all source records are roots, this is a root.
            #
            # Otherwise use the resolved global parent.
            # -----------------------------------------------------

            if parent_global_keys:

                # Remove duplicates while retaining order.
                parent_global_keys = list(
                    dict.fromkeys(
                        parent_global_keys
                    )
                )

                # In a valid tree there should normally be only
                # one physical parent.
                global_record[
                    "parent_global_key"
                ] = parent_global_keys[0]

            else:

                global_record[
                    "parent_global_key"
                ] = None

        # =====================================================
        # STEP 4
        #
        # Build global parent -> children relationship.
        # =====================================================

        children_by_parent = {}

        roots = []

        for global_key, record in global_nodes.items():

            parent_key = record.get(
                "parent_global_key"
            )

            if parent_key is None:

                roots.append(
                    global_key
                )

            else:

                children_by_parent.setdefault(
                    parent_key,
                    []
                ).append(
                    global_key
                )

        # =====================================================
        # SORT
        # =====================================================

        def sort_key(global_key):

            return str(
                global_nodes[
                    global_key
                ].get(
                    "name",
                    ""
                )
            ).lower()

        roots.sort(
            key=sort_key
        )

        for parent_key in children_by_parent:

            children_by_parent[
                parent_key
            ].sort(
                key=sort_key
            )

        # =====================================================
        # STEP 5
        #
        # Determine visibility.
        #
        # We don't throw hierarchy away just because a child
        # doesn't match the filter.
        # =====================================================

        def branch_matches(
            global_key,
            visited=None
        ):

            if visited is None:
                visited = set()

            if global_key in visited:
                return False

            visited.add(
                global_key
            )

            record = global_nodes[
                global_key
            ]

            # -------------------------------------------------
            # Node itself
            # -------------------------------------------------

            if self._record_matches(
                record,
                search,
                selected_type
            ):

                return True

            # -------------------------------------------------
            # Children
            # -------------------------------------------------

            for child_key in children_by_parent.get(
                global_key,
                []
            ):

                if branch_matches(
                    child_key,
                    visited
                ):

                    return True

            return False

        # =====================================================
        # STEP 6
        #
        # Create the visible tree.
        # =====================================================

        for global_key in roots:

            if not branch_matches(
                global_key
            ):
                continue

            self._add_global_node(
                parent_item=None,
                global_key=global_key,
                global_nodes=global_nodes,
                children_by_parent=children_by_parent,
                all_components=all_components,
                search=search,
                selected_type=selected_type,
                path=set()
            )

        # =====================================================
        # START COLLAPSED
        # =====================================================

        self.collapse_all()

    def _get_global_node_key(
        self,
        record
    ):
        """
        Returns the identity used to MERGE project copies
        of the same physical asset.

        Priority:

            1. asset_id
            2. asset tag
            3. name + type

        IMPORTANT:

            node_id is deliberately NOT used.

        node_id identifies a project-local tree node.

        asset_id identifies the physical/global asset.
        """

        asset_id = record.get(
            "asset_id"
        )

        if asset_id:

            return (
                "ASSET",
                str(
                    asset_id
                ).strip().upper()
            )

        # -----------------------------------------------------
        # Try asset tag.
        # -----------------------------------------------------

        node = record.get(
            "node"
        )

        asset_tag = ""

        if node is not None:

            asset_tag = getattr(
                node,
                "asset_tag",
                ""
            )

        if asset_tag:

            return (
                "TAG",
                str(
                    asset_tag
                ).strip().upper()
            )

        # -----------------------------------------------------
        # Final fallback.
        #
        # Name + type.
        #
        # This is weaker than asset_id and should only be used
        # when no physical identity exists.
        # -----------------------------------------------------

        name = str(
            record.get(
                "name",
                ""
            )
        ).strip().upper()

        node_type = str(
            record.get(
                "type",
                ""
            )
        ).strip().upper()

        return (
            "NAME",
            node_type,
            name
        )

    def _add_node(
        self,
        parent_item,
        record,
        node_by_id,
        project_components,
        search,
        selected_type,
        path
    ):

        node = record.get(
            "node"
        )

        if node is None:
            return

        node_id = record.get(
            "node_id"
        )

        if not node_id:
            return

        # =====================================================
        # CYCLE PROTECTION
        #
        # A corrupt assets.json should not make the UI recurse
        # into the abyss.
        # =====================================================

        if node_id in path:
            return

        path = set(
            path
        )

        path.add(
            node_id
        )

        # =====================================================
        # FIND CHILDREN USING parent_id
        #
        # THIS IS THE ONLY RELATIONSHIP USED TO BUILD THE TREE.
        #
        # asset_id is NOT used here.
        # =====================================================

        children = []

        for candidate in node_by_id.values():

            if candidate is record:
                continue

            if candidate.get(
                "parent_id"
            ) == node_id:

                children.append(
                    candidate
                )

        children.sort(
            key=lambda record:
                str(
                    record.get(
                        "name",
                        ""
                    )
                ).lower()
        )

        # =====================================================
        # CHECK WHETHER THIS NODE MATCHES THE FILTER
        # =====================================================

        direct_match = self._record_matches(
            record,
            search,
            selected_type
        )

        # =====================================================
        # CHECK WHETHER ANY DESCENDANT MATCHES
        #
        # This allows a search for P-02 to still display:
        #
        # REF-3 SS-1
        #     HV-201A
        #         P-02
        #
        # instead of ripping the hierarchy apart like a
        # badly maintained switchyard.
        # =====================================================

        descendant_match = False

        for child in children:

            if self._branch_contains_match(
                child,
                node_by_id,
                search,
                selected_type,
                set()
            ):

                descendant_match = True

                break

        # =====================================================
        # COMPONENT MATCH
        # =====================================================

        component_match = False

        node_type = self._normalise_type(
            getattr(
                node,
                "node_type",
                ""
            )
        )

        if node_type == "PANEL":

            panel_node_id = node_id

            for component_record in project_components:

                if component_record.get(
                    "panel_node_id"
                ) != panel_node_id:

                    continue

                if self._record_matches(
                    component_record,
                    search,
                    selected_type
                ):

                    component_match = True

                    break

        # =====================================================
        # SHOULD THIS NODE BE DISPLAYED?
        # =====================================================

        if (
            not direct_match
            and
            not descendant_match
            and
            not component_match
        ):

            return

        # =====================================================
        # CREATE TREE ITEM
        # =====================================================

        item = QTreeWidgetItem()

        item.setText(
            0,
            str(
                record.get(
                    "name",
                    ""
                )
            )
        )

        item.setText(
            1,
            str(
                record.get(
                    "type",
                    ""
                )
            )
        )

        item.setText(
            2,
            self._get_asset_tag(
                record
            )
        )

        # -----------------------------------------------------
        # Store complete record.
        # -----------------------------------------------------

        item.setData(
            0,
            Qt.ItemDataRole.UserRole,
            record
        )

        # =====================================================
        # ADD TO PARENT
        # =====================================================

        if parent_item is None:

            self.tree.addTopLevelItem(
                item
            )

        else:

            parent_item.addChild(
                item
            )

        # =====================================================
        # ADD CHILDREN
        # =====================================================

        for child in children:

            self._add_node(
                parent_item=item,
                record=child,
                node_by_id=node_by_id,
                project_components=project_components,
                search=search,
                selected_type=selected_type,
                path=path
            )

        # =====================================================
        # ADD COMPONENTS UNDER PANEL
        # =====================================================

        if node_type == "PANEL":

            for component_record in sorted(
                project_components,
                key=lambda record:
                    str(
                        record.get(
                            "name",
                            ""
                        )
                    ).lower()
            ):

                if component_record.get(
                    "panel_node_id"
                ) != node_id:

                    continue

                # -------------------------------------------------
                # When searching, only show matching components.
                # -------------------------------------------------

                if (
                    search
                    and
                    not self._record_matches(
                        component_record,
                        search,
                        selected_type
                    )
                ):

                    continue

                self._add_component(
                    item,
                    component_record
                )

    def _branch_contains_match(
        self,
        record,
        node_by_id,
        search,
        selected_type,
        visited
    ):

        node_id = record.get(
            "node_id"
        )

        if not node_id:
            return False

        if node_id in visited:
            return False

        visited = set(
            visited
        )

        visited.add(
            node_id
        )

        # -----------------------------------------------------
        # Does this node itself match?
        # -----------------------------------------------------

        if self._record_matches(
            record,
            search,
            selected_type
        ):

            return True

        # -----------------------------------------------------
        # Check children.
        # -----------------------------------------------------

        for child in node_by_id.values():

            if child.get(
                "parent_id"
            ) != node_id:

                continue

            if self._branch_contains_match(
                child,
                node_by_id,
                search,
                selected_type,
                visited
            ):

                return True

        return False
    # =========================================================
    # COMPONENT
    # =========================================================

    def _add_component(
        self,
        parent_item,
        record
    ):

        component = (
            record[
                "component"
            ]
        )

        item = QTreeWidgetItem()

        item.setText(
            0,
            str(
                getattr(
                    component,
                    "name",
                    ""
                )
            )
        )

        item.setText(
            1,
            str(
                record.get(
                    "type",
                    ""
                )
            )
        )

        item.setText(
            2,
            ""
        )

        item.setData(
            0,
            Qt.ItemDataRole.UserRole,
            record
        )

        parent_item.addChild(
            item
        )

    # =========================================================
    # MASTER PARENT LOOKUP
    # =========================================================

    def _get_master_parent_asset_id(
        self,
        record
    ):

        """
        Get:

            asset.metadata["parent_asset_id"]

        from the global asset library.

        This allows:

            HV-201C
                asset_id = ASSET-C...

                metadata:
                    parent_asset_id = ASSET-A17...

        to be placed below:

            REF-3 SS-1
                asset_id = ASSET-A17...
        """

        asset_id = str(
            record.get(
                "asset_id",
                ""
            )
            or ""
        ).strip()

        if not asset_id:

            return ""

        manager = (
            record.get(
                "asset_manager"
            )
        )

        if manager is None:

            return ""

        try:

            library = getattr(
                manager,
                "asset_library",
                None
            )

            if library is None:

                return ""

            try:

                library.load()

            except Exception:

                pass

            asset = (
                library.get_asset(
                    asset_id
                )
            )

            if not asset:

                return ""

            metadata = (
                asset.get(
                    "metadata"
                )
                or {}
            )

            return str(
                metadata.get(
                    "parent_asset_id",
                    ""
                )
                or ""
            ).strip()

        except Exception:

            return ""

    # =========================================================
    # FILTER MATCH
    # =========================================================

    def _record_matches(
        self,
        record,
        search,
        selected_type
    ):

        record_type = record.get(
            "record_type"
        )

        asset_type = self._normalise_type(
            record.get(
                "type",
                ""
            )
        )

        # =====================================================
        # TYPE FILTER
        # =====================================================

        if selected_type == "Substations":

            if (
                record_type != "NODE"
                or
                asset_type != "SUBSTATION"
            ):

                return False

        elif selected_type == "Switchboards":

            if (
                record_type != "NODE"
                or
                asset_type != "SWITCHBOARD"
            ):

                return False

        elif selected_type == "Panels":

            if (
                record_type != "NODE"
                or
                asset_type != "PANEL"
            ):

                return False

        elif selected_type == "Components":

            if record_type != "COMPONENT":

                return False

        # =====================================================
        # NO SEARCH TEXT
        # =====================================================

        if not search:

            return True

        # =====================================================
        # BUILD SEARCHABLE TEXT
        # =====================================================

        values = []

        node = record.get(
            "node"
        )

        component = record.get(
            "component"
        )

        # -----------------------------------------------------
        # NODE
        # -----------------------------------------------------

        if node is not None:

            fields = (

                "name",
                "node_id",
                "asset_id",
                "asset_tag",

                "manufacturer",
                "model",
                "serial_number",

                "equipment_name",
                "equipment_type",

            )

            for field in fields:

                values.append(
                    str(
                        getattr(
                            node,
                            field,
                            ""
                        )
                        or ""
                    )
                )

        # -----------------------------------------------------
        # COMPONENT
        # -----------------------------------------------------

        if component is not None:

            fields = (

                "name",
                "component_id",
                "component_type",

                "manufacturer",
                "model",
                "serial_number",

                "description",

                "ct_primary",
                "ct_secondary",
                "ct_ratio",
                "ct_class",
                "burden",
                "core",

                "vt_ratio",
                "firmware",

                "coil_voltage",
                "contact_configuration",

                "meter_type",
                "accuracy_class",

                "protection_functions",

            )

            for field in fields:

                values.append(
                    str(
                        getattr(
                            component,
                            field,
                            ""
                        )
                        or ""
                    )
                )

        # -----------------------------------------------------
        # PROJECT
        # -----------------------------------------------------

        values.append(
            str(
                record.get(
                    "project",
                    ""
                )
            )
        )

        searchable = " ".join(
            values
        ).lower()

        return (
            search
            in searchable
        )
    # =========================================================
    # SELECTION
    # =========================================================

    def _selection_changed(self):

        item = (
            self.tree.currentItem()
        )

        if item is None:

            self._clear_details()

            return

        record = item.data(
            0,
            Qt.ItemDataRole.UserRole
        )

        if not isinstance(
            record,
            dict
        ):

            self._clear_details()

            return

        self._show_record(
            record
        )

    # =========================================================
    # SHOW RECORD
    # =========================================================

    def _show_record(
        self,
        record
    ):

        self._clear_detail_widgets()

        if (
            record.get(
                "record_type"
            )
            == "COMPONENT"
        ):

            component = (
                record[
                    "component"
                ]
            )

            self.details_title.setText(
                str(
                    getattr(
                        component,
                        "name",
                        "Component"
                    )
                )
            )

            self.details_subtitle.setText(
                (
                    f"{record.get('type', '')} | "
                    f"Projects: "
                    f"{self._occurrence_projects(record)}"
                )
            )

            self._show_component_details(
                record
            )

            self._show_test_history(
                record
            )

            return

        node = (
            record[
                "node"
            ]
        )

        self.details_title.setText(
            str(
                getattr(
                    node,
                    "name",
                    "Asset"
                )
            )
        )

        self.details_subtitle.setText(
            (
                f"{record.get('type', '')} | "
                f"Asset ID: "
                f"{getattr(node, 'asset_id', '') or 'Not assigned'} | "
                f"Projects: "
                f"{self._occurrence_projects(record)}"
            )
        )

        self._show_node_details(
            record
        )

        if (
            self._normalise_type(
                getattr(
                    node,
                    "node_type",
                    ""
                )
            )
            == "PANEL"
        ):

            self._show_panel_components(
                record
            )

            self._show_test_history(
                record
            )

    # =========================================================
    # NODE DETAILS
    # =========================================================

    def _show_node_details(
        self,
        record
    ):

        node = (
            record[
                "node"
            ]
        )

        self._add_section(
            "Physical Asset"
        )

        fields = [

            (
                "Asset ID",
                getattr(
                    node,
                    "asset_id",
                    ""
                )
            ),

            (
                "Asset Tag",
                self._get_asset_tag(
                    record
                )
            ),

            (
                "Name",
                getattr(
                    node,
                    "name",
                    ""
                )
            ),

            (
                "Asset Type",
                getattr(
                    node,
                    "node_type",
                    ""
                )
            ),

            (
                "Manufacturer",
                getattr(
                    node,
                    "manufacturer",
                    ""
                )
            ),

            (
                "Model",
                getattr(
                    node,
                    "model",
                    ""
                )
            ),

            (
                "Serial Number",
                getattr(
                    node,
                    "serial_number",
                    ""
                )
            ),

            (
                "Feed Equipment",
                getattr(
                    node,
                    "equipment_name",
                    ""
                )
            ),

            (
                "Equipment Type",
                getattr(
                    node,
                    "equipment_type",
                    ""
                )
            ),

        ]

        if (
            self._normalise_type(
                getattr(
                    node,
                    "node_type",
                    ""
                )
            )
            == "PANEL"
        ):

            fields.extend(

                [

                    (
                        "Number of CTs",
                        getattr(
                            node,
                            "ct_count",
                            0
                        )
                    ),

                    (
                        "Numerical Relays",
                        getattr(
                            node,
                            "relay_count",
                            0
                        )
                    ),

                    (
                        "Auxiliary Relays",
                        getattr(
                            node,
                            "aux_count",
                            0
                        )
                    ),

                    (
                        "Meters",
                        getattr(
                            node,
                            "meter_count",
                            0
                        )
                    ),

                ]

            )

        for label, value in fields:

            self._add_field(
                label,
                value
            )

        self._add_section(
            "Project Occurrences"
        )

        for occurrence in (
            record.get(
                "occurrences",
                []
            )
        ):

            self._add_field(
                "Project",
                (
                    f"{occurrence.get('project', '')} | "
                    f"Node: "
                    f"{occurrence.get('node_id', '-')}"
                )
            )

    # =========================================================
    # COMPONENT DETAILS
    # =========================================================

    def _show_component_details(
        self,
        record
    ):

        component = (
            record[
                "component"
            ]
        )

        self._add_section(
            "Component Configuration"
        )

        fields = [

            (
                "Component ID",
                getattr(
                    component,
                    "component_id",
                    ""
                )
            ),

            (
                "Component Type",
                getattr(
                    component,
                    "component_type",
                    ""
                )
            ),

            (
                "Manufacturer",
                getattr(
                    component,
                    "manufacturer",
                    ""
                )
            ),

            (
                "Model",
                getattr(
                    component,
                    "model",
                    ""
                )
            ),

            (
                "Serial Number",
                getattr(
                    component,
                    "serial_number",
                    ""
                )
            ),

            (
                "Description",
                getattr(
                    component,
                    "description",
                    ""
                )
            ),

            (
                "CT Primary",
                getattr(
                    component,
                    "ct_primary",
                    ""
                )
            ),

            (
                "CT Secondary",
                getattr(
                    component,
                    "ct_secondary",
                    ""
                )
            ),

            (
                "CT Ratio",
                getattr(
                    component,
                    "ct_ratio",
                    ""
                )
            ),

            (
                "CT Class",
                getattr(
                    component,
                    "ct_class",
                    ""
                )
            ),

            (
                "Burden",
                getattr(
                    component,
                    "burden",
                    ""
                )
            ),

            (
                "Core",
                getattr(
                    component,
                    "core",
                    ""
                )
            ),

            (
                "VT Ratio",
                getattr(
                    component,
                    "vt_ratio",
                    ""
                )
            ),

            (
                "Firmware",
                getattr(
                    component,
                    "firmware",
                    ""
                )
            ),

            (
                "Coil Voltage",
                getattr(
                    component,
                    "coil_voltage",
                    ""
                )
            ),

            (
                "Contact Configuration",
                getattr(
                    component,
                    "contact_configuration",
                    ""
                )
            ),

            (
                "Meter Type",
                getattr(
                    component,
                    "meter_type",
                    ""
                )
            ),

            (
                "Meter Functions",
                self._format_value(
                    getattr(
                        component,
                        "meter_functions",
                        ""
                    )
                )
            ),

            (
                "Accuracy Class",
                getattr(
                    component,
                    "accuracy_class",
                    ""
                )
            ),

            (
                "Protection Functions",
                self._format_value(
                    getattr(
                        component,
                        "protection_functions",
                        ""
                    )
                )
            ),

        ]

        for label, value in fields:

            self._add_field(
                label,
                value
            )

    # =========================================================
    # PANEL COMPONENTS
    # =========================================================

    def _show_panel_components(
        self,
        panel_record
    ):

        panel = (
            panel_record[
                "node"
            ]
        )

        panel_asset_id = str(
            getattr(
                panel,
                "asset_id",
                ""
            )
            or ""
        ).strip()

        panel_node_ids = set(
            panel_record.get(
                "node_ids",
                []
            )
        )

        current_panel_id = getattr(
            panel,
            "node_id",
            None
        )

        if current_panel_id:

            panel_node_ids.add(
                current_panel_id
            )

        components = []

        for record in self._all_records:

            if (
                record.get(
                    "record_type"
                )
                != "COMPONENT"
            ):

                continue

            component_panel = (
                record.get(
                    "node"
                )
            )

            if component_panel is None:

                continue

            component_panel_asset_id = str(
                getattr(
                    component_panel,
                    "asset_id",
                    ""
                )
                or ""
            ).strip()

            component_panel_node_id = getattr(
                component_panel,
                "node_id",
                None
            )

            if (

                (
                    panel_asset_id
                    and
                    component_panel_asset_id
                    and
                    panel_asset_id
                    ==
                    component_panel_asset_id
                )

                or

                (
                    component_panel_node_id
                    in
                    panel_node_ids
                )

            ):

                components.append(
                    record
                )

        self._add_section(
            "Configured Components"
        )

        if not components:

            self._add_field(
                "Components",
                "No configured components found."
            )

            return

        table = QTableWidget()

        table.setColumnCount(
            5
        )

        table.setHorizontalHeaderLabels(
            [
                "Component",
                "Type",
                "Manufacturer",
                "Model",
                "Serial Number",
            ]
        )

        table.setRowCount(
            len(
                components
            )
        )

        for row, record in enumerate(
            components
        ):

            component = (
                record[
                    "component"
                ]
            )

            values = [

                getattr(
                    component,
                    "name",
                    ""
                ),

                getattr(
                    component,
                    "component_type",
                    ""
                ),

                getattr(
                    component,
                    "manufacturer",
                    ""
                ),

                getattr(
                    component,
                    "model",
                    ""
                ),

                getattr(
                    component,
                    "serial_number",
                    ""
                ),

            ]

            for column, value in enumerate(
                values
            ):

                table.setItem(
                    row,
                    column,
                    QTableWidgetItem(
                        str(
                            value
                            or ""
                        )
                    )
                )

        table.setEditTriggers(
            QTableWidget.EditTrigger.NoEditTriggers
        )

        table.setSelectionBehavior(
            QTableWidget.SelectionBehavior.SelectRows
        )

        table.horizontalHeader().setSectionResizeMode(
            QHeaderView.ResizeMode.ResizeToContents
        )

        table.setMaximumHeight(
            250
        )

        self.details_form_layout.addWidget(
            table
        )

    # =========================================================
    # TEST HISTORY
    # =========================================================

    def _show_test_history(
        self,
        record
    ):

        self._add_section(
            "Test History"
        )

        rows = (
            self._get_test_history(
                record
            )
        )

        table = QTableWidget()

        table.setColumnCount(
            8
        )

        table.setHorizontalHeaderLabels(
            [
                "Date",
                "Project",
                "Panel",
                "Test Type",
                "Protection / Component",
                "Result",
                "Remarks",
                "Test ID",
            ]
        )

        table.setRowCount(
            len(
                rows
            )
        )

        for row_index, row_data in enumerate(
            rows
        ):

            for column_index, value in enumerate(
                row_data
            ):

                table.setItem(
                    row_index,
                    column_index,
                    QTableWidgetItem(
                        str(
                            value
                            or ""
                        )
                    )
                )

        table.setEditTriggers(
            QTableWidget.EditTrigger.NoEditTriggers
        )

        table.setSelectionBehavior(
            QTableWidget.SelectionBehavior.SelectRows
        )

        table.horizontalHeader().setSectionResizeMode(
            QHeaderView.ResizeMode.ResizeToContents
        )

        table.setMinimumHeight(
            180
        )

        self.details_form_layout.addWidget(
            table
        )

        if not rows:

            self._add_field(
                "Tests",
                "No test history found"
            )

    def _get_test_history(
        self,
        record
    ):

        results = []

        node = (
            record.get(
                "node"
            )
        )

        component = (
            record.get(
                "component"
            )
        )

        if node is None:

            return results

        panel_id = getattr(
            node,
            "node_id",
            None
        )

        component_id = None

        if component is not None:

            component_id = getattr(
                component,
                "component_id",
                None
            )

        occurrences = (
            record.get(
                "occurrences",
                []
            )
        )

        if not occurrences:

            occurrences = [

                {

                    "project":
                        record.get(
                            "project",
                            ""
                        ),

                    "folder":
                        record.get(
                            "folder"
                        ),

                    "node":
                        node,

                    "node_id":
                        panel_id,

                }

            ]

        seen = set()

        for occurrence in occurrences:

            folder = occurrence.get(
                "folder"
            )

            if not folder:

                continue

            folder = Path(
                folder
            )

            database_candidates = [

                folder / "testing.db",

                folder / "tests.db",

            ]

            database_file = next(
                (
                    path

                    for path in database_candidates

                    if path.exists()
                ),
                None
            )

            if database_file is None:

                continue

            project = (
                occurrence.get(
                    "project",
                    ""
                )
            )

            try:

                connection = sqlite3.connect(
                    str(
                        database_file
                    )
                )

                cursor = connection.cursor()

                component_names = (
                    self._load_component_names(
                        folder
                    )
                )

                if component_id is not None:

                    try:

                        cursor.execute(
                            """
                            SELECT
                                test_id,
                                test_date,
                                component_id,
                                test_type,
                                result,
                                remarks
                            FROM component_tests
                            WHERE component_id = ?
                            ORDER BY test_date DESC
                            """,
                            (
                                component_id,
                            )
                        )

                        for row in cursor.fetchall():

                            result_row = [

                                row[1],

                                project,

                                getattr(
                                    node,
                                    "name",
                                    ""
                                ),

                                "COMPONENT TEST",

                                component_names.get(
                                    row[2],
                                    row[2]
                                    or ""
                                ),

                                row[4],

                                row[5],

                                row[0],

                            ]

                            key = tuple(
                                str(
                                    value
                                    or ""
                                )
                                for value
                                in result_row
                            )

                            if key not in seen:

                                seen.add(
                                    key
                                )

                                results.append(
                                    result_row
                                )

                    except sqlite3.Error:

                        pass

                else:

                    try:

                        cursor.execute(
                            """
                            SELECT
                                test_id,
                                test_date,
                                component_id,
                                test_type,
                                result,
                                remarks
                            FROM component_tests
                            WHERE panel_id = ?
                            ORDER BY test_date DESC
                            """,
                            (
                                panel_id,
                            )
                        )

                        for row in cursor.fetchall():

                            result_row = [

                                row[1],

                                project,

                                getattr(
                                    node,
                                    "name",
                                    ""
                                ),

                                "COMPONENT TEST",

                                component_names.get(
                                    row[2],
                                    row[2]
                                    or ""
                                ),

                                row[4],

                                row[5],

                                row[0],

                            ]

                            key = tuple(
                                str(
                                    value
                                    or ""
                                )
                                for value
                                in result_row
                            )

                            if key not in seen:

                                seen.add(
                                    key
                                )

                                results.append(
                                    result_row
                                )

                    except sqlite3.Error:

                        pass

                connection.close()

            except sqlite3.Error:

                continue

        results.sort(
            key=lambda row:
                str(
                    row[0]
                    or ""
                ),
            reverse=True
        )

        return results

    # =========================================================
    # COMPONENT NAMES
    # =========================================================

    @staticmethod
    def _load_component_names(
        project_folder
    ):

        names = {}

        components_file = (
            Path(
                project_folder
            )
            /
            "components.json"
        )

        if not components_file.exists():

            return names

        try:

            import json

            with open(
                components_file,
                "r",
                encoding="utf-8"
            ) as file:

                data = json.load(
                    file
                )

            if isinstance(
                data,
                list
            ):

                for item in data:

                    if not isinstance(
                        item,
                        dict
                    ):

                        continue

                    component_id = (
                        item.get(
                            "component_id"
                        )
                    )

                    if component_id:

                        names[
                            component_id
                        ] = item.get(
                            "name",
                            ""
                        )

        except Exception:

            pass

        return names

    # =========================================================
    # DETAILS
    # =========================================================

    def _clear_details(self):

        self.details_title.setText(
            "Select an asset"
        )

        self.details_subtitle.setText(
            ""
        )

        self._clear_detail_widgets()

    def _clear_detail_widgets(self):

        while (
            self.details_form_layout.count()
        ):

            item = (
                self.details_form_layout
                .takeAt(0)
            )

            widget = item.widget()

            if widget is not None:

                widget.deleteLater()

    def _add_section(
        self,
        title
    ):

        label = QLabel(
            title
        )

        label.setObjectName(
            "Section"
        )

        self.details_form_layout.addWidget(
            label
        )

    def _add_field(
        self,
        label,
        value
    ):

        row = QHBoxLayout()

        name = QLabel(
            str(
                label
            )
        )

        name.setMinimumWidth(
            170
        )

        name.setStyleSheet(
            "font-weight: 600;"
        )

        value_label = QLabel(
            self._format_value(
                value
            )
        )

        value_label.setObjectName(
            "Value"
        )

        value_label.setWordWrap(
            True
        )

        row.addWidget(
            name
        )

        row.addWidget(
            value_label,
            1
        )

        self.details_form_layout.addLayout(
            row
        )

    # =========================================================
    # EXPANSION
    # =========================================================

    def expand_all(self):

        self.tree.expandAll()

    def collapse_all(self):

        self.tree.collapseAll()

    # =========================================================
    # PHYSICAL KEY
    # =========================================================

    @staticmethod
    def _physical_key(
        record
    ):

        asset_id = str(
            record.get(
                "asset_id",
                ""
            )
            or ""
        ).strip()

        if asset_id:

            return (
                f"ASSET::{asset_id}"
            )

        node = (
            record.get(
                "node"
            )
        )

        node_id = getattr(
            node,
            "node_id",
            None
        )

        return (
            f"NODE::{node_id}"
        )

    # =========================================================
    # PROJECTS
    # =========================================================

    @staticmethod
    def _occurrence_projects(
        record
    ):

        projects = []

        for occurrence in (
            record.get(
                "occurrences",
                []
            )
        ):

            project = str(
                occurrence.get(
                    "project",
                    ""
                )
            ).strip()

            if (
                project
                and
                project not in projects
            ):

                projects.append(
                    project
                )

        if not projects:

            project = str(
                record.get(
                    "project",
                    ""
                )
            ).strip()

            if project:

                projects.append(
                    project
                )

        return ", ".join(
            projects
        )

    # =========================================================
    # ASSET TAG
    # =========================================================

    def _get_asset_tag(
        self,
        record
    ):

        node = (
            record.get(
                "node"
            )
        )

        if node is None:

            return ""

        value = getattr(
            node,
            "asset_tag",
            ""
        )

        if value:

            return str(
                value
            )

        manager = (
            record.get(
                "asset_manager"
            )
        )

        asset_id = getattr(
            node,
            "asset_id",
            None
        )

        if (
            manager is not None
            and
            asset_id
        ):

            try:

                library = getattr(
                    manager,
                    "asset_library",
                    None
                )

                if library is not None:

                    try:

                        library.load()

                    except Exception:

                        pass

                    asset = (
                        library.get_asset(
                            asset_id
                        )
                    )

                    if asset:

                        return str(
                            asset.get(
                                "asset_tag",
                                ""
                            )
                        )

            except Exception:

                pass

        return ""

    # =========================================================
    # FORMAT
    # =========================================================

    @staticmethod
    def _format_value(
        value
    ):

        if value is None:

            return ""

        if isinstance(
            value,
            (
                list,
                tuple
            )
        ):

            return ", ".join(
                str(
                    item
                )
                for item in value
            )

        if isinstance(
            value,
            dict
        ):

            return ", ".join(
                f"{key}: {value}"
                for key, value
                in value.items()
            )

        return str(
            value
        )

    # =========================================================
    # NORMALISE
    # =========================================================

    @staticmethod
    def _normalise_type(
        value
    ):

        value = str(
            value
            or ""
        ).strip().upper()

        return (
            value
            .replace(
                "-",
                "_"
            )
            .replace(
                " ",
                "_"
            )
        )

    @staticmethod
    def _normalise_component_type(
        value
    ):

        value = str(
            value
            or ""
        ).strip().upper()

        aliases = {

            "CURRENT TRANSFORMER":
                "CT",

            "CURRENT_TRANSFORMER":
                "CT",

            "CT":
                "CT",

            "NUMERICAL RELAY":
                "NUMERICAL RELAY",

            "NUMERICAL_RELAY":
                "NUMERICAL RELAY",

            "RELAY":
                "NUMERICAL RELAY",

            "AUXILIARY RELAY":
                "AUXILIARY RELAY",

            "AUXILIARY_RELAY":
                "AUXILIARY RELAY",

            "AUX RELAY":
                "AUXILIARY RELAY",

            "METER":
                "METER",

            "AMMETER":
                "METER",

            "VOLTMETER":
                "METER",

            "MULTIFUNCTION METER":
                "METER",

            "MULTIFUNCTION_METER":
                "METER",

        }

        return aliases.get(
            value,
            value.replace(
                "_",
                " "
            )
        )

    # =========================================================
    # EXPORT
    # =========================================================

    def export_asset_register(
        self
    ):

        default_path = (
            PROJECTS_DIR
            /
            "Asset_Register.xlsx"
        )

        output_path, _ = (
            QFileDialog.getSaveFileName(
                self,
                "Export Asset Register",
                str(
                    default_path
                ),
                "Excel Workbook (*.xlsx)"
            )
        )

        if not output_path:

            return

        try:

            self._export_excel(
                output_path
            )

            QMessageBox.information(
                self,
                "Asset Register",
                (
                    "Asset Register exported successfully.\n\n"
                    f"{output_path}"
                )
            )

        except PermissionError:

            QMessageBox.critical(
                self,
                "Export Failed",
                (
                    "The Excel file is currently open or "
                    "you do not have permission to overwrite it."
                )
            )

        except Exception as error:

            QMessageBox.critical(
                self,
                "Export Failed",
                str(
                    error
                )
            )

    # =========================================================
    # EXCEL
    # =========================================================

    def _export_excel(
        self,
        output_path
    ):

        self.global_asset_service.refresh()

        self._load_records()

        workbook = Workbook()

        workbook.remove(
            workbook.active
        )

        columns = [

            "Project",
            "Substation",
            "Switchboard",
            "Panel",
            "Panel Asset Tag",
            "Feed Equipment",
            "Equipment Type",

            "Component",
            "Component Type",
            "Component ID",

            "Manufacturer",
            "Model",
            "Serial Number",
            "Description",

            "CT Primary",
            "CT Secondary",
            "CT Ratio",
            "CT Class",
            "Burden",
            "Core",

            "VT Ratio",
            "Firmware",

            "Coil Voltage",
            "Contact Configuration",

            "Meter Type",
            "Meter Functions",
            "Accuracy Class",

            "Protection Functions",

        ]

        rows = []

        for record in self._all_records:

            if (
                record.get(
                    "record_type"
                )
                != "COMPONENT"
            ):

                continue

            component = (
                record[
                    "component"
                ]
            )

            panel = (
                record[
                    "node"
                ]
            )

            hierarchy = (
                self._get_hierarchy(
                    record
                )
            )

            rows.append(
                {

                    "Project":
                        record.get(
                            "project",
                            ""
                        ),

                    "Substation":
                        hierarchy.get(
                            "SUBSTATION",
                            ""
                        ),

                    "Switchboard":
                        hierarchy.get(
                            "SWITCHBOARD",
                            ""
                        ),

                    "Panel":
                        getattr(
                            panel,
                            "name",
                            ""
                        ),

                    "Panel Asset Tag":
                        self._get_asset_tag(
                            record
                        ),

                    "Feed Equipment":
                        getattr(
                            panel,
                            "equipment_name",
                            ""
                        ),

                    "Equipment Type":
                        getattr(
                            panel,
                            "equipment_type",
                            ""
                        ),

                    "Component":
                        getattr(
                            component,
                            "name",
                            ""
                        ),

                    "Component Type":
                        getattr(
                            component,
                            "component_type",
                            ""
                        ),

                    "Component ID":
                        getattr(
                            component,
                            "component_id",
                            ""
                        ),

                    "Manufacturer":
                        getattr(
                            component,
                            "manufacturer",
                            ""
                        ),

                    "Model":
                        getattr(
                            component,
                            "model",
                            ""
                        ),

                    "Serial Number":
                        getattr(
                            component,
                            "serial_number",
                            ""
                        ),

                    "Description":
                        getattr(
                            component,
                            "description",
                            ""
                        ),

                    "CT Primary":
                        getattr(
                            component,
                            "ct_primary",
                            ""
                        ),

                    "CT Secondary":
                        getattr(
                            component,
                            "ct_secondary",
                            ""
                        ),

                    "CT Ratio":
                        getattr(
                            component,
                            "ct_ratio",
                            ""
                        ),

                    "CT Class":
                        getattr(
                            component,
                            "ct_class",
                            ""
                        ),

                    "Burden":
                        getattr(
                            component,
                            "burden",
                            ""
                        ),

                    "Core":
                        getattr(
                            component,
                            "core",
                            ""
                        ),

                    "VT Ratio":
                        getattr(
                            component,
                            "vt_ratio",
                            ""
                        ),

                    "Firmware":
                        getattr(
                            component,
                            "firmware",
                            ""
                        ),

                    "Coil Voltage":
                        getattr(
                            component,
                            "coil_voltage",
                            ""
                        ),

                    "Contact Configuration":
                        getattr(
                            component,
                            "contact_configuration",
                            ""
                        ),

                    "Meter Type":
                        getattr(
                            component,
                            "meter_type",
                            ""
                        ),

                    "Meter Functions":
                        self._format_value(
                            getattr(
                                component,
                                "meter_functions",
                                ""
                            )
                        ),

                    "Accuracy Class":
                        getattr(
                            component,
                            "accuracy_class",
                            ""
                        ),

                    "Protection Functions":
                        self._format_value(
                            getattr(
                                component,
                                "protection_functions",
                                ""
                            )
                        ),
                }
            )

        self._write_sheet(
            workbook.create_sheet(
                "Asset Register"
            ),
            columns,
            rows,
            "AssetRegisterTable"
        )

        # -----------------------------------------------------
        # PANELS
        # -----------------------------------------------------

        panel_columns = [

            "Project",
            "Substation",
            "Switchboard",
            "Panel",
            "Panel Asset Tag",
            "Panel ID",
            "Feed Equipment",
            "Equipment Type",
            "CT Count",
            "Numerical Relay Count",
            "Auxiliary Relay Count",
            "Meter Count",

        ]

        panel_rows = []

        for record in self._all_records:

            if (
                record.get(
                    "record_type"
                )
                != "NODE"
            ):

                continue

            node = record[
                "node"
            ]

            if (
                self._normalise_type(
                    getattr(
                        node,
                        "node_type",
                        ""
                    )
                )
                != "PANEL"
            ):

                continue

            hierarchy = (
                self._get_hierarchy(
                    record
                )
            )

            panel_rows.append(
                {

                    "Project":
                        record.get(
                            "project",
                            ""
                        ),

                    "Substation":
                        hierarchy.get(
                            "SUBSTATION",
                            ""
                        ),

                    "Switchboard":
                        hierarchy.get(
                            "SWITCHBOARD",
                            ""
                        ),

                    "Panel":
                        getattr(
                            node,
                            "name",
                            ""
                        ),

                    "Panel Asset Tag":
                        self._get_asset_tag(
                            record
                        ),

                    "Panel ID":
                        getattr(
                            node,
                            "node_id",
                            ""
                        ),

                    "Feed Equipment":
                        getattr(
                            node,
                            "equipment_name",
                            ""
                        ),

                    "Equipment Type":
                        getattr(
                            node,
                            "equipment_type",
                            ""
                        ),

                    "CT Count":
                        getattr(
                            node,
                            "ct_count",
                            0
                        ),

                    "Numerical Relay Count":
                        getattr(
                            node,
                            "relay_count",
                            0
                        ),

                    "Auxiliary Relay Count":
                        getattr(
                            node,
                            "aux_count",
                            0
                        ),

                    "Meter Count":
                        getattr(
                            node,
                            "meter_count",
                            0
                        ),
                }
            )

        self._write_sheet(
            workbook.create_sheet(
                "Panels"
            ),
            panel_columns,
            panel_rows,
            "PanelsTable"
        )

        # -----------------------------------------------------
        # INFO
        # -----------------------------------------------------

        info = workbook.create_sheet(
            "Register Info"
        )

        info["A1"] = (
            "Protection Testing Suite"
        )

        info["A1"].font = Font(
            bold=True,
            size=16
        )

        info["A3"] = (
            "Generated"
        )

        info["B3"] = (
            datetime.now().strftime(
                "%Y-%m-%d %H:%M:%S"
            )
        )

        info["A4"] = (
            "Projects"
        )

        info["B4"] = len(
            self.global_asset_service
            .get_projects()
        )

        info["A5"] = (
            "Physical Assets"
        )

        info["B5"] = len(
            [
                record

                for record in self._all_records

                if record.get(
                    "record_type"
                )
                == "NODE"
            ]
        )

        info["A6"] = (
            "Components"
        )

        info["B6"] = len(
            [
                record

                for record in self._all_records

                if record.get(
                    "record_type"
                )
                == "COMPONENT"
            ]
        )

        info.column_dimensions[
            "A"
        ].width = 25

        info.column_dimensions[
            "B"
        ].width = 30

        workbook.save(
            output_path
        )

    # =========================================================
    # EXCEL WRITER
    # =========================================================

    @staticmethod
    def _write_sheet(
        worksheet,
        columns,
        rows,
        table_name
    ):

        for column_index, column in enumerate(
            columns,
            start=1
        ):

            cell = worksheet.cell(
                row=1,
                column=column_index,
                value=column
            )

            cell.font = Font(
                bold=True
            )

            cell.alignment = Alignment(
                horizontal="center",
                vertical="center"
            )

        for row_index, row in enumerate(
            rows,
            start=2
        ):

            for column_index, column in enumerate(
                columns,
                start=1
            ):

                worksheet.cell(
                    row=row_index,
                    column=column_index,
                    value=row.get(
                        column,
                        ""
                    )
                )

        worksheet.freeze_panes = (
            "A2"
        )

        if rows:

            last_column = (
                get_column_letter(
                    len(columns)
                )
            )

            table = Table(
                displayName=table_name,
                ref=(
                    f"A1:"
                    f"{last_column}"
                    f"{len(rows) + 1}"
                )
            )

            table.tableStyleInfo = (
                TableStyleInfo(
                    name="TableStyleMedium2",
                    showFirstColumn=False,
                    showLastColumn=False,
                    showRowStripes=True,
                    showColumnStripes=False,
                )
            )

            worksheet.add_table(
                table
            )

        for column_index in range(
            1,
            len(columns) + 1
        ):

            letter = (
                get_column_letter(
                    column_index
                )
            )

            maximum = len(
                str(
                    worksheet.cell(
                        row=1,
                        column=column_index
                    ).value
                    or ""
                )
            )

            for row_index in range(
                2,
                min(
                    worksheet.max_row,
                    1000
                ) + 1
            ):

                value = worksheet.cell(
                    row=row_index,
                    column=column_index
                ).value

                if value is not None:

                    maximum = max(
                        maximum,
                        len(
                            str(
                                value
                            )
                        )
                    )

            worksheet.column_dimensions[
                letter
            ].width = min(
                max(
                    maximum + 2,
                    12
                ),
                45
            )

    # =========================================================
    # HIERARCHY FOR EXPORT
    # =========================================================

    def _get_hierarchy(
        self,
        record
    ):

        result = {}

        current = record

        visited = set()

        while current is not None:

            physical_key = (
                self._physical_key(
                    current
                )
            )

            if physical_key in visited:

                break

            visited.add(
                physical_key
            )

            node = (
                current.get(
                    "node"
                )
            )

            if node is None:

                break

            node_type = (
                self._normalise_type(
                    getattr(
                        node,
                        "node_type",
                        ""
                    )
                )
            )

            result[
                node_type
            ] = getattr(
                node,
                "name",
                ""
            )

            current = (
                self._find_parent_record(
                    current
                )
            )

        return result

    def _find_parent_record(
        self,
        child_record
    ):

        node_records = [

            record

            for record in self._all_records

            if record.get(
                "record_type"
            )
            == "NODE"

        ]

        child_project = (
            child_record.get(
                "project",
                ""
            )
        )

        # -----------------------------------------------------
        # First: project-local parent_id.
        # -----------------------------------------------------

        for occurrence in (
            child_record.get(
                "occurrences",
                []
            )
        ):

            parent_id = (
                occurrence.get(
                    "parent_id"
                )
            )

            if not parent_id:

                continue

            project = (
                occurrence.get(
                    "project",
                    child_project
                )
            )

            for candidate in node_records:

                for candidate_occurrence in (
                    candidate.get(
                        "occurrences",
                        []
                    )
                ):

                    if (

                        candidate_occurrence.get(
                            "project"
                        )
                        == project

                        and

                        candidate_occurrence.get(
                            "node_id"
                        )
                        == parent_id

                    ):

                        return candidate

        # -----------------------------------------------------
        # Second: global parent_asset_id.
        # -----------------------------------------------------

        master_parent_id = (
            self._get_master_parent_asset_id(
                child_record
            )
        )

        if master_parent_id:

            for candidate in node_records:

                candidate_asset_id = str(
                    candidate.get(
                        "asset_id",
                        ""
                    )
                    or ""
                ).strip()

                if (
                    candidate_asset_id
                    == master_parent_id
                ):

                    return candidate

        return None
    def _add_global_node(
        self,
        parent_item,
        global_key,
        global_nodes,
        children_by_parent,
        all_components,
        search,
        selected_type,
        path
    ):

        # =====================================================
        # CYCLE PROTECTION
        # =====================================================

        if global_key in path:
            return

        path = set(
            path
        )

        path.add(
            global_key
        )

        record = global_nodes[
            global_key
        ]

        # =====================================================
        # MATCH
        # =====================================================

        direct_match = self._record_matches(
            record,
            search,
            selected_type
        )

        descendant_match = False

        for child_key in children_by_parent.get(
            global_key,
            []
        ):

            if self._global_branch_matches(
                child_key,
                global_nodes,
                children_by_parent,
                search,
                selected_type,
                set()
            ):

                descendant_match = True

                break

        # =====================================================
        # COMPONENT MATCH
        # =====================================================

        component_match = False

        node = record.get(
            "node"
        )

        node_type = self._normalise_type(
            record.get(
                "type",
                ""
            )
        )

        if node_type == "PANEL":

            # -------------------------------------------------
            # A physical panel can have component records from
            # multiple projects.
            #
            # Match them using the project-local node identity
            # stored in panel_node_id.
            # -------------------------------------------------

            source_records = record.get(
                "records",
                []
            )

            panel_node_keys = set()

            for source_record in source_records:

                project = str(
                    source_record.get(
                        "project",
                        ""
                    )
                )

                node_id = source_record.get(
                    "node_id"
                )

                if node_id:

                    panel_node_keys.add(
                        (
                            project,
                            node_id
                        )
                    )

            for component_record in all_components:

                component_project = str(
                    component_record.get(
                        "project",
                        ""
                    )
                )

                panel_node_id = component_record.get(
                    "panel_node_id"
                )

                if (
                    component_project,
                    panel_node_id
                ) not in panel_node_keys:

                    continue

                if self._record_matches(
                    component_record,
                    search,
                    selected_type
                ):

                    component_match = True

                    break

        # =====================================================
        # FILTER
        # =====================================================

        if (
            not direct_match
            and
            not descendant_match
            and
            not component_match
        ):

            return

        # =====================================================
        # CREATE ITEM
        # =====================================================

        item = QTreeWidgetItem()

        item.setText(
            0,
            str(
                record.get(
                    "name",
                    ""
                )
            )
        )

        item.setText(
            1,
            str(
                record.get(
                    "type",
                    ""
                )
            )
        )

        item.setText(
            2,
            self._get_asset_tag(
                record
            )
        )

        # -----------------------------------------------------
        # Store the GLOBAL record.
        # -----------------------------------------------------

        item.setData(
            0,
            Qt.ItemDataRole.UserRole,
            record
        )

        # =====================================================
        # ADD TO TREE
        # =====================================================

        if parent_item is None:

            self.tree.addTopLevelItem(
                item
            )

        else:

            parent_item.addChild(
                item
            )

        # =====================================================
        # STORE TREE NODE
        # =====================================================

        self._tree_nodes[
            global_key
        ] = item

        # =====================================================
        # CHILDREN
        # =====================================================

        for child_key in children_by_parent.get(
            global_key,
            []
        ):

            self._add_global_node(
                parent_item=item,
                global_key=child_key,
                global_nodes=global_nodes,
                children_by_parent=children_by_parent,
                all_components=all_components,
                search=search,
                selected_type=selected_type,
                path=path
            )

        # =====================================================
        # COMPONENTS
        # =====================================================

        if node_type == "PANEL":

            source_records = record.get(
                "records",
                []
            )

            panel_node_keys = set()

            for source_record in source_records:

                project = str(
                    source_record.get(
                        "project",
                        ""
                    )
                )

                node_id = source_record.get(
                    "node_id"
                )

                if node_id:

                    panel_node_keys.add(
                        (
                            project,
                            node_id
                        )
                    )

            matching_components = []

            for component_record in all_components:

                key = (
                    str(
                        component_record.get(
                            "project",
                            ""
                        )
                    ),
                    component_record.get(
                        "panel_node_id"
                    )
                )

                if key in panel_node_keys:

                    if (
                        not search
                        or
                        self._record_matches(
                            component_record,
                            search,
                            selected_type
                        )
                    ):

                        matching_components.append(
                            component_record
                        )

            matching_components.sort(
                key=lambda r:
                    str(
                        r.get(
                            "name",
                            ""
                        )
                    ).lower()
            )

            for component_record in matching_components:

                self._add_component(
                    item,
                    component_record
                )

    def _global_branch_matches(
        self,
        global_key,
        global_nodes,
        children_by_parent,
        search,
        selected_type,
        visited
    ):

        if global_key in visited:
            return False

        visited.add(
            global_key
        )

        record = global_nodes[
            global_key
        ]

        if self._record_matches(
            record,
            search,
            selected_type
        ):

            return True

        for child_key in children_by_parent.get(
            global_key,
            []
        ):

            if self._global_branch_matches(
                child_key,
                global_nodes,
                children_by_parent,
                search,
                selected_type,
                visited
            ):

                return True

        return False