from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import (
    Font,
    PatternFill,
    Alignment,
    Border,
    Side,
)
from openpyxl.utils import get_column_letter

from app.services.asset_library_manager import (
    AssetLibraryManager,
)


class AssetExportService:
    """
    Global asset register exporter.

    Source of truth:
        Global AssetLibraryManager

    Project selection is NOT required.

    Workbook contains:

        Summary
        Substations
        Switchboards
        Panels
        Components
        CT Register
        Numerical Relay Register
        Auxiliary Relay Register
        Meter Register
    """

    # =========================================================
    # INITIALIZATION
    # =========================================================

    def __init__(
        self,
        asset_library=None,
    ):

        self.asset_library = (
            asset_library
            if asset_library is not None
            else AssetLibraryManager()
        )

    # =========================================================
    # PUBLIC EXPORT
    # =========================================================

    def export_asset_register(
        self,
        output_path,
    ):

        output_path = Path(
            output_path
        )

        # -----------------------------------------------------
        # Reload global asset database
        # -----------------------------------------------------

        try:

            self.asset_library.load()

        except Exception:

            pass

        assets = (
            self.asset_library.get_all_assets()
            or []
        )

        # -----------------------------------------------------
        # Separate physical assets
        # -----------------------------------------------------

        substations = []
        switchboards = []
        panels = []

        for asset in assets:

            if not isinstance(
                asset,
                dict,
            ):
                continue

            asset_type = self._type(
                asset
            )

            if asset_type == "SUBSTATION":

                substations.append(
                    asset
                )

            elif asset_type == "SWITCHBOARD":

                switchboards.append(
                    asset
                )

            elif asset_type == "PANEL":

                panels.append(
                    asset
                )

        # -----------------------------------------------------
        # Build hierarchy
        # -----------------------------------------------------

        hierarchy = (
            self._build_hierarchy(
                substations,
                switchboards,
                panels,
            )
        )

        # -----------------------------------------------------
        # Workbook
        # -----------------------------------------------------

        workbook = Workbook()

        default_sheet = (
            workbook.active
        )

        workbook.remove(
            default_sheet
        )

        # -----------------------------------------------------
        # Summary
        # -----------------------------------------------------

        self._write_summary(
            workbook,
            substations,
            switchboards,
            panels,
        )

        # -----------------------------------------------------
        # Substations
        # -----------------------------------------------------

        self._write_substations(
            workbook,
            substations,
        )

        # -----------------------------------------------------
        # Switchboards
        # -----------------------------------------------------

        self._write_switchboards(
            workbook,
            switchboards,
            hierarchy,
        )

        # -----------------------------------------------------
        # Panels
        # -----------------------------------------------------

        self._write_panels(
            workbook,
            panels,
            hierarchy,
        )

        # -----------------------------------------------------
        # Components
        # -----------------------------------------------------

        components = (
            self._collect_components(
                panels,
                hierarchy,
            )
        )

        self._write_components(
            workbook,
            components,
        )

        # -----------------------------------------------------
        # Component type sheets
        # -----------------------------------------------------

        self._write_component_type_sheet(
            workbook,
            "CT Register",
            components,
            (
                "CT",
                "CURRENT TRANSFORMER",
            ),
        )

        self._write_component_type_sheet(
            workbook,
            "Numerical Relay Register",
            components,
            (
                "NUMERICAL_RELAY",
                "NUMERICAL RELAY",
            ),
        )

        self._write_component_type_sheet(
            workbook,
            "Auxiliary Relay Register",
            components,
            (
                "AUXILIARY_RELAY",
                "AUX RELAY",
                "AUXILIARY RELAY",
            ),
        )

        self._write_component_type_sheet(
            workbook,
            "Meter Register",
            components,
            (
                "METER",
                "AMMETER",
                "VOLTMETER",
                "MULTIFUNCTION_METER",
                "MULTIFUNCTION METER",
            ),
        )

        # -----------------------------------------------------
        # Save
        # -----------------------------------------------------

        output_path.parent.mkdir(
            parents=True,
            exist_ok=True,
        )

        workbook.save(
            output_path
        )

        return output_path

    # =========================================================
    # HIERARCHY
    # =========================================================

    def _build_hierarchy(
        self,
        substations,
        switchboards,
        panels,
    ):

        hierarchy = {
            "substations": {},
            "switchboards": {},
            "panels": {},
        }

        # -----------------------------------------------------
        # Substations
        # -----------------------------------------------------

        for substation in substations:

            asset_id = self._clean_value(
                substation.get(
                    "asset_id"
                )
            )

            if not asset_id:
                continue

            hierarchy[
                "substations"
            ][
                asset_id
            ] = substation

        # -----------------------------------------------------
        # Switchboards
        # -----------------------------------------------------

        for switchboard in switchboards:

            metadata = (
                switchboard.get(
                    "metadata"
                )
                or {}
            )

            parent_id = self._clean_value(
                metadata.get(
                    "parent_asset_id"
                )
            )

            asset_id = self._clean_value(
                switchboard.get(
                    "asset_id"
                )
            )

            if not asset_id:
                continue

            hierarchy[
                "switchboards"
            ][
                asset_id
            ] = {
                "asset":
                    switchboard,

                "parent_id":
                    parent_id,
            }

        # -----------------------------------------------------
        # Panels
        # -----------------------------------------------------

        for panel in panels:

            metadata = (
                panel.get(
                    "metadata"
                )
                or {}
            )

            parent_id = self._clean_value(
                metadata.get(
                    "parent_asset_id"
                )
            )

            asset_id = self._clean_value(
                panel.get(
                    "asset_id"
                )
            )

            if not asset_id:
                continue

            hierarchy[
                "panels"
            ][
                asset_id
            ] = {
                "asset":
                    panel,

                "parent_id":
                    parent_id,
            }

        return hierarchy

    # =========================================================
    # COMPONENT COLLECTION
    # =========================================================

    def _collect_components(
        self,
        panels,
        hierarchy,
    ):

        components = []

        for panel in panels:

            panel_id = self._clean_value(
                panel.get(
                    "asset_id"
                )
            )

            metadata = (
                panel.get(
                    "metadata"
                )
                or {}
            )

            snapshots = (
                metadata.get(
                    "components"
                )
                or []
            )

            switchboard_id = (
                hierarchy[
                    "panels"
                ]
                .get(
                    panel_id,
                    {},
                )
                .get(
                    "parent_id"
                )
            )

            switchboard_info = (
                hierarchy[
                    "switchboards"
                ]
                .get(
                    switchboard_id,
                    {},
                )
            )

            switchboard = (
                switchboard_info.get(
                    "asset"
                )
                or {}
            )

            substation_id = (
                switchboard_info.get(
                    "parent_id"
                )
            )

            substation = (
                hierarchy[
                    "substations"
                ]
                .get(
                    substation_id,
                    {},
                )
                or {}
            )

            for component in snapshots:

                if not isinstance(
                    component,
                    dict,
                ):
                    continue

                components.append(
                    {
                        "substation":
                            substation,

                        "switchboard":
                            switchboard,

                        "panel":
                            panel,

                        "component":
                            component,
                    }
                )

        return components

    # =========================================================
    # SUMMARY
    # =========================================================

    def _write_summary(
        self,
        workbook,
        substations,
        switchboards,
        panels,
    ):

        sheet = (
            self._new_sheet(
                workbook,
                "Summary",
                [
                    "Asset Type",
                    "Count",
                ],
            )
        )

        rows = [
            (
                "Substations",
                len(substations),
            ),
            (
                "Switchboards",
                len(switchboards),
            ),
            (
                "Panels",
                len(panels),
            ),
        ]

        for row in rows:

            sheet.append(
                row
            )

        self._finish_sheet(
            sheet
        )

    # =========================================================
    # SUBSTATIONS
    # =========================================================

    def _write_substations(
        self,
        workbook,
        substations,
    ):

        headers = [
            "Asset ID",
            "Substation",
            "Asset Tag",
            "Manufacturer",
            "Model",
            "Serial Number",
        ]

        sheet = (
            self._new_sheet(
                workbook,
                "Substations",
                headers,
            )
        )

        for asset in substations:

            sheet.append(
                [
                    self._clean_value(
                        asset.get(
                            "asset_id"
                        )
                    ),

                    self._clean_value(
                        asset.get(
                            "name"
                        )
                    ),

                    self._clean_value(
                        asset.get(
                            "asset_tag"
                        )
                    ),

                    self._clean_value(
                        asset.get(
                            "manufacturer"
                        )
                    ),

                    self._clean_value(
                        asset.get(
                            "model"
                        )
                    ),

                    self._clean_value(
                        asset.get(
                            "serial_number"
                        )
                    ),
                ]
            )

        self._finish_sheet(
            sheet
        )

    # =========================================================
    # SWITCHBOARDS
    # =========================================================

    def _write_switchboards(
        self,
        workbook,
        switchboards,
        hierarchy,
    ):

        headers = [
            "Asset ID",
            "Substation",
            "Switchboard",
            "Asset Tag",
            "Manufacturer",
            "Model",
            "Serial Number",
        ]

        sheet = (
            self._new_sheet(
                workbook,
                "Switchboards",
                headers,
            )
        )

        for asset in switchboards:

            metadata = (
                asset.get(
                    "metadata"
                )
                or {}
            )

            parent_id = self._clean_value(
                metadata.get(
                    "parent_asset_id"
                )
            )

            parent = (
                hierarchy[
                    "substations"
                ]
                .get(
                    parent_id,
                    {},
                )
                or {}
            )

            sheet.append(
                [
                    self._clean_value(
                        asset.get(
                            "asset_id"
                        )
                    ),

                    self._clean_value(
                        parent.get(
                            "name"
                        )
                    ),

                    self._clean_value(
                        asset.get(
                            "name"
                        )
                    ),

                    self._clean_value(
                        asset.get(
                            "asset_tag"
                        )
                    ),

                    self._clean_value(
                        asset.get(
                            "manufacturer"
                        )
                    ),

                    self._clean_value(
                        asset.get(
                            "model"
                        )
                    ),

                    self._clean_value(
                        asset.get(
                            "serial_number"
                        )
                    ),
                ]
            )

        self._finish_sheet(
            sheet
        )

    # =========================================================
    # PANELS
    # =========================================================

    def _write_panels(
        self,
        workbook,
        panels,
        hierarchy,
    ):

        headers = [
            "Asset ID",
            "Substation",
            "Switchboard",
            "Panel",
            "Asset Tag",
            "Feed Equipment",
            "Equipment Type",
            "CT Count",
            "Numerical Relay Count",
            "Auxiliary Relay Count",
            "Meter Count",
            "Manufacturer",
            "Model",
            "Serial Number",
        ]

        sheet = (
            self._new_sheet(
                workbook,
                "Panels",
                headers,
            )
        )

        for panel in panels:

            metadata = (
                panel.get(
                    "metadata"
                )
                or {}
            )

            switchboard_id = self._clean_value(
                metadata.get(
                    "parent_asset_id"
                )
            )

            switchboard_info = (
                hierarchy[
                    "switchboards"
                ]
                .get(
                    switchboard_id,
                    {},
                )
            )

            switchboard = (
                switchboard_info.get(
                    "asset"
                )
                or {}
            )

            substation_id = (
                switchboard_info.get(
                    "parent_id"
                )
            )

            substation = (
                hierarchy[
                    "substations"
                ]
                .get(
                    substation_id,
                    {},
                )
                or {}
            )

            components = (
                metadata.get(
                    "components"
                )
                or []
            )

            counts = (
                self._component_counts(
                    components
                )
            )

            sheet.append(
                [
                    self._clean_value(
                        panel.get(
                            "asset_id"
                        )
                    ),

                    self._clean_value(
                        substation.get(
                            "name"
                        )
                    ),

                    self._clean_value(
                        switchboard.get(
                            "name"
                        )
                    ),

                    self._clean_value(
                        panel.get(
                            "name"
                        )
                    ),

                    self._clean_value(
                        panel.get(
                            "asset_tag"
                        )
                    ),

                    self._clean_value(
                        panel.get(
                            "equipment_name"
                        )
                    ),

                    self._clean_value(
                        panel.get(
                            "equipment_type"
                        )
                    ),

                    counts["cts"],

                    counts["relays"],

                    counts["aux"],

                    counts["meters"],

                    self._clean_value(
                        panel.get(
                            "manufacturer"
                        )
                    ),

                    self._clean_value(
                        panel.get(
                            "model"
                        )
                    ),

                    self._clean_value(
                        panel.get(
                            "serial_number"
                        )
                    ),
                ]
            )

        self._finish_sheet(
            sheet
        )

    # =========================================================
    # ALL COMPONENTS
    # =========================================================

    def _write_components(
        self,
        workbook,
        components,
    ):

        headers = self._component_headers()

        sheet = (
            self._new_sheet(
                workbook,
                "Components",
                headers,
            )
        )

        for record in components:

            self._append_component(
                sheet,
                record,
            )

        self._finish_sheet(
            sheet
        )

    # =========================================================
    # COMPONENT TYPE SHEETS
    # =========================================================

    def _write_component_type_sheet(
        self,
        workbook,
        title,
        components,
        accepted_types,
    ):

        accepted_types = {
            str(value)
            .strip()
            .upper()
            for value in accepted_types
        }

        headers = self._component_headers()

        sheet = (
            self._new_sheet(
                workbook,
                title,
                headers,
            )
        )

        for record in components:

            component = (
                record.get(
                    "component"
                )
                or {}
            )

            component_type = self._type(
                component
            )

            if component_type not in accepted_types:

                continue

            self._append_component(
                sheet,
                record,
            )

        self._finish_sheet(
            sheet
        )

    # =========================================================
    # COMPONENT HEADERS
    # =========================================================

    @staticmethod
    def _component_headers():

        return [
            "Substation",
            "Switchboard",
            "Panel",
            "Panel Asset Tag",

            "Component",
            "Component Type",

            "Manufacturer",
            "Model",
            "Serial Number",
            "Description",

            "CT Primary",
            "CT Secondary",
            "CT Ratio",
            "CT Class",
            "Burden",
            "Core",

            "VT Ratio",
            "Firmware",

            "Coil Voltage",
            "Contact Configuration",

            "Meter Type",
            "Meter Functions",
            "Accuracy Class",

            "Protection Functions",
        ]

    # =========================================================
    # COMPONENT ROW
    # =========================================================

    def _append_component(
        self,
        sheet,
        record,
    ):

        substation = (
            record.get(
                "substation"
            )
            or {}
        )

        switchboard = (
            record.get(
                "switchboard"
            )
            or {}
        )

        panel = (
            record.get(
                "panel"
            )
            or {}
        )

        component = (
            record.get(
                "component"
            )
            or {}
        )

        # -----------------------------------------------------
        # Component type
        # -----------------------------------------------------

        component_type = self._clean_value(
            component.get(
                "component_type",
                component.get(
                    "type",
                    "",
                )
            )
        )

        # -----------------------------------------------------
        # Protection functions
        # -----------------------------------------------------

        protection_functions = (
            component.get(
                "protection_functions",
                []
            )
            or []
        )

        # -----------------------------------------------------
        # Meter functions
        # -----------------------------------------------------

        meter_functions = (
            component.get(
                "meter_functions",
                []
            )
            or []
        )

        # -----------------------------------------------------
        # Build row
        #
        # IMPORTANT:
        # This order MUST exactly match
        # _component_headers().
        # -----------------------------------------------------

        row = [

            # 1
            self._clean_value(
                substation.get(
                    "name"
                )
            ),

            # 2
            self._clean_value(
                switchboard.get(
                    "name"
                )
            ),

            # 3
            self._clean_value(
                panel.get(
                    "name"
                )
            ),

            # 4
            self._clean_value(
                panel.get(
                    "asset_tag"
                )
            ),

            # 5
            self._clean_value(
                component.get(
                    "name"
                )
            ),

            # 6
            component_type,

            # 7
            self._clean_value(
                component.get(
                    "manufacturer"
                )
            ),

            # 8
            self._clean_value(
                component.get(
                    "model"
                )
            ),

            # 9
            self._clean_value(
                component.get(
                    "serial_number"
                )
            ),

            # 10
            self._clean_value(
                component.get(
                    "description"
                )
            ),

            # 11
            self._clean_value(
                component.get(
                    "ct_primary"
                )
            ),

            # 12
            self._clean_value(
                component.get(
                    "ct_secondary"
                )
            ),

            # 13
            self._clean_value(
                component.get(
                    "ct_ratio"
                )
            ),

            # 14
            self._clean_value(
                component.get(
                    "ct_class"
                )
            ),

            # 15
            self._clean_value(
                component.get(
                    "burden"
                )
            ),

            # 16
            self._clean_value(
                component.get(
                    "core"
                )
            ),

            # 17
            self._clean_value(
                component.get(
                    "vt_ratio"
                )
            ),

            # 18
            self._clean_value(
                component.get(
                    "firmware"
                )
            ),

            # 19
            self._clean_value(
                component.get(
                    "coil_voltage"
                )
            ),

            # 20
            self._clean_value(
                component.get(
                    "contact_configuration"
                )
            ),

            # 21
            self._clean_value(
                component.get(
                    "meter_type"
                )
            ),

            # 22
            self._list_to_text(
                meter_functions
            ),

            # 23
            self._clean_value(
                component.get(
                    "accuracy_class"
                )
            ),

            # 24
            self._list_to_text(
                protection_functions
            ),
        ]

        # -----------------------------------------------------
        # Safety check
        # -----------------------------------------------------

        expected_columns = len(
            self._component_headers()
        )

        if len(row) != expected_columns:

            raise ValueError(
                "Component export column mismatch: "
                f"expected {expected_columns}, "
                f"got {len(row)} "
                f"for component "
                f"{component.get('name', '')!r}"
            )

        sheet.append(
            row
        )

    # =========================================================
    # COMPONENT COUNTS
    # =========================================================

    def _component_counts(
        self,
        components,
    ):

        counts = {
            "cts": 0,
            "relays": 0,
            "aux": 0,
            "meters": 0,
        }

        for component in components:

            if not isinstance(
                component,
                dict,
            ):
                continue

            component_type = self._type(
                component
            )

            if component_type in (
                "CT",
                "CURRENT TRANSFORMER",
            ):

                counts["cts"] += 1

            elif component_type in (
                "NUMERICAL_RELAY",
                "NUMERICAL RELAY",
            ):

                counts["relays"] += 1

            elif component_type in (
                "AUXILIARY_RELAY",
                "AUX RELAY",
                "AUXILIARY RELAY",
            ):

                counts["aux"] += 1

            elif component_type in (
                "METER",
                "AMMETER",
                "VOLTMETER",
                "MULTIFUNCTION_METER",
                "MULTIFUNCTION METER",
            ):

                counts["meters"] += 1

        return counts

    # =========================================================
    # TYPE
    # =========================================================

    @staticmethod
    def _type(
        asset
    ):

        if not isinstance(
            asset,
            dict,
        ):

            return ""

        value = asset.get(
            "component_type"
        )

        if value is None:

            value = asset.get(
                "asset_type",
                ""
            )

        return str(
            AssetExportService._clean_value(
                value
            )
        ).strip().upper()

    # =========================================================
    # LIST TO TEXT
    # =========================================================

    @staticmethod
    def _list_to_text(
        value
    ):

        if value is None:

            return ""

        if isinstance(
            value,
            (list, tuple, set)
        ):

            cleaned = []

            for item in value:

                item = (
                    AssetExportService
                    ._clean_value(
                        item
                    )
                )

                if item != "":

                    cleaned.append(
                        str(item)
                    )

            return ", ".join(
                cleaned
            )

        return str(
            AssetExportService
            ._clean_value(
                value
            )
        )

    # =========================================================
    # CLEAN VALUE
    # =========================================================

    @staticmethod
    def _clean_value(
        value
    ):
        """
        Convert values into safe Excel-friendly values.

        Particularly handles accidental one-element tuples
        such as:

            ('P-02',)

        which previously appeared in the exported workbook.
        """

        if value is None:

            return ""

        # -----------------------------------------------------
        # Tuple / list / set
        # -----------------------------------------------------

        if isinstance(
            value,
            (tuple, list, set)
        ):

            values = list(
                value
            )

            if not values:

                return ""

            # Single-value container:
            # ('P-02',) -> P-02
            if len(values) == 1:

                return (
                    AssetExportService
                    ._clean_value(
                        values[0]
                    )
                )

            return ", ".join(
                str(
                    AssetExportService
                    ._clean_value(
                        item
                    )
                )
                for item in values
            )

        # -----------------------------------------------------
        # Strings
        # -----------------------------------------------------

        if isinstance(
            value,
            str,
        ):

            return value.strip()

        # -----------------------------------------------------
        # Numbers / other values
        # -----------------------------------------------------

        return value

    # =========================================================
    # CREATE SHEET
    # =========================================================

    def _new_sheet(
        self,
        workbook,
        title,
        headers,
    ):

        sheet = workbook.create_sheet(
            title
        )

        for column, header in enumerate(
            headers,
            start=1,
        ):

            cell = sheet.cell(
                row=1,
                column=column,
                value=header,
            )

            cell.font = Font(
                bold=True,
                color="FFFFFF",
            )

            cell.fill = PatternFill(
                "solid",
                fgColor="333333",
            )

            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=True,
            )

        sheet.freeze_panes = "A2"

        sheet.auto_filter.ref = (
            f"A1:"
            f"{get_column_letter(len(headers))}"
            f"1"
        )

        sheet.row_dimensions[
            1
        ].height = 30

        return sheet

    # =========================================================
    # FINISH SHEET
    # =========================================================

    def _finish_sheet(
        self,
        sheet,
    ):

        thin = Side(
            style="thin",
            color="CCCCCC",
        )

        border = Border(
            bottom=thin
        )

        for row in sheet.iter_rows():

            for cell in row:

                cell.border = border

                cell.alignment = Alignment(
                    vertical="top",
                    wrap_text=True,
                )

        # -----------------------------------------------------
        # Column widths
        # -----------------------------------------------------

        for column_cells in sheet.columns:

            if not column_cells:

                continue

            column_letter = (
                get_column_letter(
                    column_cells[0].column
                )
            )

            maximum = 0

            for cell in column_cells:

                value = str(
                    cell.value
                    or ""
                )

                maximum = max(
                    maximum,
                    len(value),
                )

            sheet.column_dimensions[
                column_letter
            ].width = min(
                max(
                    maximum + 2,
                    12,
                ),
                35,
            )