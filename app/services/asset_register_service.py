from pathlib import Path
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


class AssetRegisterService:

    """
    Generates the master Excel Asset Register from all projects.

    The project JSON/configuration remains the source of truth.
    Excel is only the generated engineering register.
    """

    # =========================================================
    # COMPONENT FIELDS
    # =========================================================

    COMPONENT_FIELDS = [

        "manufacturer",
        "model",
        "serial_number",
        "description",

        # CT
        "ct_primary",
        "ct_secondary",
        "ct_ratio",
        "ct_class",
        "burden",
        "core",

        # Numerical relay
        "vt_ratio",
        "firmware",

        # Auxiliary relay
        "coil_voltage",
        "contact_configuration",

        # Meter
        "meter_type",
        "meter_functions",
        "accuracy_class",

        # Protection
        "protection_functions",
    ]

    # =========================================================
    # MASTER REGISTER COLUMNS
    # =========================================================

    MASTER_COLUMNS = [

        "Project",

        "Substation",

        "Switchboard",

        "Panel",

        "Panel Asset Tag",

        "Feed Equipment",

        "Equipment Type",

        "Panel Component",

        "Component Type",

        "Component ID",

        "Manufacturer",

        "Model",

        "Serial Number",

        "Description",

        "CT Primary",

        "CT Secondary",

        "CT Ratio",

        "CT Class",

        "Burden",

        "Core",

        "VT Ratio",

        "Firmware",

        "Coil Voltage",

        "Contact Configuration",

        "Meter Type",

        "Meter Functions",

        "Accuracy Class",

        "Protection Functions",

    ]

    # =========================================================
    # INIT
    # =========================================================

    def __init__(
        self,
        global_asset_service
    ):

        self.global_asset_service = (
            global_asset_service
        )

    # =========================================================
    # EXPORT
    # =========================================================

    def export(
        self,
        output_path
    ):

        output_path = Path(
            output_path
        )

        output_path.parent.mkdir(
            parents=True,
            exist_ok=True
        )

        self.global_asset_service.refresh()

        workbook = Workbook()

        # Remove default sheet
        default_sheet = workbook.active

        workbook.remove(
            default_sheet
        )

        # =====================================================
        # MASTER REGISTER
        # =====================================================

        master_sheet = workbook.create_sheet(
            "Asset Register"
        )

        master_rows = (
            self.build_master_rows()
        )

        self.write_sheet(
            master_sheet,
            self.MASTER_COLUMNS,
            master_rows,
            "AssetRegisterTable"
        )

        # =====================================================
        # PANELS
        # =====================================================

        panel_sheet = workbook.create_sheet(
            "Panels"
        )

        panel_columns = [

            "Project",
            "Substation",
            "Switchboard",
            "Panel",
            "Panel Asset Tag",
            "Panel ID",
            "Feed Equipment",
            "Equipment Type",
            "CT Count",
            "Numerical Relay Count",
            "Auxiliary Relay Count",
            "Meter Count",
        ]

        panel_rows = (
            self.build_panel_rows()
        )

        self.write_sheet(
            panel_sheet,
            panel_columns,
            panel_rows,
            "PanelsTable"
        )

        # =====================================================
        # CT
        # =====================================================

        ct_sheet = workbook.create_sheet(
            "CTs"
        )

        ct_columns = self.get_component_columns()

        ct_rows = (
            self.build_component_rows(
                "CT"
            )
        )

        self.write_sheet(
            ct_sheet,
            ct_columns,
            ct_rows,
            "CTTable"
        )

        # =====================================================
        # NUMERICAL RELAYS
        # =====================================================

        relay_sheet = workbook.create_sheet(
            "Numerical Relays"
        )

        relay_rows = (
            self.build_component_rows(
                "NUMERICAL_RELAY"
            )
        )

        self.write_sheet(
            relay_sheet,
            ct_columns,
            relay_rows,
            "NumericalRelayTable"
        )

        # =====================================================
        # AUX RELAYS
        # =====================================================

        aux_sheet = workbook.create_sheet(
            "Aux Relays"
        )

        aux_rows = (
            self.build_component_rows(
                "AUXILIARY_RELAY"
            )
        )

        self.write_sheet(
            aux_sheet,
            ct_columns,
            aux_rows,
            "AuxRelayTable"
        )

        # =====================================================
        # METERS
        # =====================================================

        meter_sheet = workbook.create_sheet(
            "Meters"
        )

        meter_rows = (
            self.build_component_rows(
                "METER"
            )
        )

        self.write_sheet(
            meter_sheet,
            ct_columns,
            meter_rows,
            "MeterTable"
        )

        # =====================================================
        # EXPORT INFORMATION
        # =====================================================

        info_sheet = workbook.create_sheet(
            "Register Info"
        )

        info_sheet["A1"] = (
            "Protection Testing Suite"
        )

        info_sheet["A1"].font = Font(
            bold=True,
            size=16
        )

        info_sheet["A3"] = (
            "Register Generated"
        )

        info_sheet["B3"] = (
            datetime.now().strftime(
                "%Y-%m-%d %H:%M:%S"
            )
        )

        info_sheet["A4"] = (
            "Projects Included"
        )

        info_sheet["B4"] = len(
            self.global_asset_service
            .get_projects()
        )

        info_sheet["A5"] = (
            "Total Components"
        )

        info_sheet["B5"] = len(
            self.global_asset_service
            .get_all_components()
        )

        info_sheet.column_dimensions[
            "A"
        ].width = 25

        info_sheet.column_dimensions[
            "B"
        ].width = 30

        workbook.save(
            output_path
        )

        return output_path

    # =========================================================
    # MASTER ROWS
    # =========================================================

    def build_master_rows(self):

        rows = []

        for entry in (
            self.global_asset_service
            .get_all_components()
        ):

            project_name = (
                entry["project"]
            )

            panel = (
                entry["panel"]
            )

            component = (
                entry["component"]
            )

            hierarchy = (
                self.get_hierarchy(
                    entry["asset_manager"],
                    panel
                )
            )

            asset_tag = (
                self.get_asset_tag(
                    entry["asset_manager"],
                    panel
                )
            )

            row = {

                "Project":
                    project_name,

                "Substation":
                    hierarchy.get(
                        "SUBSTATION",
                        ""
                    ),

                "Switchboard":
                    hierarchy.get(
                        "SWITCHBOARD",
                        ""
                    ),

                "Panel":
                    getattr(
                        panel,
                        "name",
                        ""
                    ),

                "Panel Asset Tag":
                    asset_tag,

                "Feed Equipment":
                    getattr(
                        panel,
                        "equipment_name",
                        ""
                    ),

                "Equipment Type":
                    getattr(
                        panel,
                        "equipment_type",
                        ""
                    ),

                "Panel Component":
                    getattr(
                        component,
                        "name",
                        ""
                    ),

                "Component Type":
                    getattr(
                        component,
                        "component_type",
                        ""
                    ),

                "Component ID":
                    getattr(
                        component,
                        "component_id",
                        ""
                    ),

            }

            self.add_component_fields(
                row,
                component
            )

            rows.append(
                row
            )

        return rows

    # =========================================================
    # PANEL ROWS
    # =========================================================

    def build_panel_rows(self):

        rows = []

        for entry in (
            self.global_asset_service
            .get_all_nodes()
        ):

            node = (
                entry["node"]
            )

            node_type = str(
                getattr(
                    node,
                    "node_type",
                    ""
                )
            ).upper()

            if node_type != "PANEL":
                continue

            asset_manager = (
                entry["asset_manager"]
            )

            hierarchy = (
                self.get_hierarchy(
                    asset_manager,
                    node
                )
            )

            row = {

                "Project":
                    entry["project"],

                "Substation":
                    hierarchy.get(
                        "SUBSTATION",
                        ""
                    ),

                "Switchboard":
                    hierarchy.get(
                        "SWITCHBOARD",
                        ""
                    ),

                "Panel":
                    getattr(
                        node,
                        "name",
                        ""
                    ),

                "Panel Asset Tag":
                    self.get_asset_tag(
                        asset_manager,
                        node
                    ),

                "Panel ID":
                    getattr(
                        node,
                        "node_id",
                        ""
                    ),

                "Feed Equipment":
                    getattr(
                        node,
                        "equipment_name",
                        ""
                    ),

                "Equipment Type":
                    getattr(
                        node,
                        "equipment_type",
                        ""
                    ),

                "CT Count":
                    getattr(
                        node,
                        "ct_count",
                        0
                    ),

                "Numerical Relay Count":
                    getattr(
                        node,
                        "relay_count",
                        0
                    ),

                "Auxiliary Relay Count":
                    getattr(
                        node,
                        "aux_count",
                        0
                    ),

                "Meter Count":
                    getattr(
                        node,
                        "meter_count",
                        0
                    ),

            }

            rows.append(
                row
            )

        return rows

    # =========================================================
    # COMPONENT ROWS
    # =========================================================

    def build_component_rows(
        self,
        requested_type
    ):

        rows = []

        requested_type = (
            str(
                requested_type
            )
            .strip()
            .upper()
        )

        for entry in (
            self.global_asset_service
            .get_all_components()
        ):

            component = (
                entry["component"]
            )

            component_type = (
                str(
                    getattr(
                        component,
                        "component_type",
                        ""
                    )
                )
                .strip()
                .upper()
            )

            if not self.component_type_matches(
                component_type,
                requested_type
            ):
                continue

            panel = (
                entry["panel"]
            )

            asset_manager = (
                entry["asset_manager"]
            )

            hierarchy = (
                self.get_hierarchy(
                    asset_manager,
                    panel
                )
            )

            row = {

                "Project":
                    entry["project"],

                "Substation":
                    hierarchy.get(
                        "SUBSTATION",
                        ""
                    ),

                "Switchboard":
                    hierarchy.get(
                        "SWITCHBOARD",
                        ""
                    ),

                "Panel":
                    getattr(
                        panel,
                        "name",
                        ""
                    ),

                "Panel Asset Tag":
                    self.get_asset_tag(
                        asset_manager,
                        panel
                    ),

                "Panel Component":
                    getattr(
                        component,
                        "name",
                        ""
                    ),

                "Component Type":
                    getattr(
                        component,
                        "component_type",
                        ""
                    ),

                "Component ID":
                    getattr(
                        component,
                        "component_id",
                        ""
                    ),
            }

            self.add_component_fields(
                row,
                component
            )

            rows.append(
                row
            )

        return rows

    # =========================================================
    # ADD COMPONENT FIELDS
    # =========================================================

    def add_component_fields(
        self,
        row,
        component
    ):

        for field in (
            self.COMPONENT_FIELDS
        ):

            value = getattr(
                component,
                field,
                ""
            )

            # Lists are converted into readable text.
            if isinstance(
                value,
                (list, tuple)
            ):

                value = ", ".join(
                    str(item)
                    for item in value
                )

            elif isinstance(
                value,
                dict
            ):

                value = str(
                    value
                )

            row[
                self.field_to_column(
                    field
                )
            ] = value

    # =========================================================
    # COMPONENT COLUMNS
    # =========================================================

    @staticmethod
    def get_component_columns():

        return [

            "Project",
            "Substation",
            "Switchboard",
            "Panel",
            "Panel Asset Tag",
            "Panel Component",
            "Component Type",
            "Component ID",

            "Manufacturer",
            "Model",
            "Serial Number",
            "Description",

            "CT Primary",
            "CT Secondary",
            "CT Ratio",
            "CT Class",
            "Burden",
            "Core",

            "VT Ratio",
            "Firmware",

            "Coil Voltage",
            "Contact Configuration",

            "Meter Type",
            "Meter Functions",
            "Accuracy Class",

            "Protection Functions",
        ]

    # =========================================================
    # FIELD → EXCEL COLUMN
    # =========================================================

    @staticmethod
    def field_to_column(
        field
    ):

        mapping = {

            "manufacturer":
                "Manufacturer",

            "model":
                "Model",

            "serial_number":
                "Serial Number",

            "description":
                "Description",

            "ct_primary":
                "CT Primary",

            "ct_secondary":
                "CT Secondary",

            "ct_ratio":
                "CT Ratio",

            "ct_class":
                "CT Class",

            "burden":
                "Burden",

            "core":
                "Core",

            "vt_ratio":
                "VT Ratio",

            "firmware":
                "Firmware",

            "coil_voltage":
                "Coil Voltage",

            "contact_configuration":
                "Contact Configuration",

            "meter_type":
                "Meter Type",

            "meter_functions":
                "Meter Functions",

            "accuracy_class":
                "Accuracy Class",

            "protection_functions":
                "Protection Functions",
        }

        return mapping.get(
            field,
            field
        )

    # =========================================================
    # COMPONENT TYPE MATCHING
    # =========================================================

    @staticmethod
    def component_type_matches(
        actual,
        requested
    ):

        actual = (
            str(actual)
            .strip()
            .upper()
        )

        requested = (
            str(requested)
            .strip()
            .upper()
        )

        if requested == "CT":

            return actual in (
                "CT",
                "CURRENT TRANSFORMER"
            )

        if requested == "NUMERICAL_RELAY":

            return actual in (
                "NUMERICAL_RELAY",
                "NUMERICAL RELAY",
                "RELAY"
            )

        if requested == "AUXILIARY_RELAY":

            return actual in (
                "AUXILIARY_RELAY",
                "AUXILIARY RELAY",
                "AUX RELAY"
            )

        if requested == "METER":

            return actual in (
                "METER",
                "AMMETER",
                "VOLTMETER",
                "MULTIFUNCTION_METER",
                "MULTIFUNCTION METER"
            )

        return actual == requested

    # =========================================================
    # HIERARCHY
    # =========================================================

    @staticmethod
    def get_hierarchy(
        asset_manager,
        node
    ):

        result = {}

        current = node

        visited = set()

        while current is not None:

            node_id = getattr(
                current,
                "node_id",
                None
            )

            if node_id in visited:
                break

            visited.add(
                node_id
            )

            node_type = str(
                getattr(
                    current,
                    "node_type",
                    ""
                )
            ).upper()

            result[
                node_type
            ] = getattr(
                current,
                "name",
                ""
            )

            parent_id = getattr(
                current,
                "parent_id",
                None
            )

            if parent_id is None:
                break

            current = (
                asset_manager.get_node(
                    parent_id
                )
            )

        return result

    # =========================================================
    # ASSET TAG
    # =========================================================

    @staticmethod
    def get_asset_tag(
        asset_manager,
        node
    ):

        asset_id = getattr(
            node,
            "asset_id",
            None
        )

        if asset_id:

            try:

                asset_manager.asset_library.load()

                asset = (
                    asset_manager
                    .asset_library
                    .get_asset(
                        asset_id
                    )
                )

                if asset:

                    return (
                        asset.get(
                            "asset_tag",
                            ""
                        )
                    )

            except Exception:
                pass

        # Fallback
        return getattr(
            node,
            "asset_tag",
            ""
        )

    # =========================================================
    # WRITE SHEET
    # =========================================================

    @staticmethod
    def write_sheet(
        worksheet,
        columns,
        rows,
        table_name
    ):

        # -----------------------------------------------------
        # HEADER
        # -----------------------------------------------------

        for column_index, column in enumerate(
            columns,
            start=1
        ):

            cell = worksheet.cell(
                row=1,
                column=column_index
            )

            cell.value = column

            cell.font = Font(
                bold=True
            )

            cell.alignment = Alignment(
                horizontal="center",
                vertical="center"
            )

        # -----------------------------------------------------
        # DATA
        # -----------------------------------------------------

        for row_index, row_data in enumerate(
            rows,
            start=2
        ):

            for column_index, column in enumerate(
                columns,
                start=1
            ):

                value = row_data.get(
                    column,
                    ""
                )

                if value is None:
                    value = ""

                worksheet.cell(
                    row=row_index,
                    column=column_index
                ).value = value

        # -----------------------------------------------------
        # FREEZE HEADER
        # -----------------------------------------------------

        worksheet.freeze_panes = "A2"

        # -----------------------------------------------------
        # FILTER / TABLE
        # -----------------------------------------------------

        last_row = max(
            1,
            len(rows) + 1
        )

        last_column = len(
            columns
        )

        if last_row >= 2:

            reference = (
                f"A1:"
                f"{get_column_letter(last_column)}"
                f"{last_row}"
            )

            table = Table(
                displayName=table_name,
                ref=reference
            )

            style = TableStyleInfo(
                name="TableStyleMedium2",
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=True,
                showColumnStripes=False
            )

            table.tableStyleInfo = style

            worksheet.add_table(
                table
            )

        # -----------------------------------------------------
        # COLUMN WIDTHS
        # -----------------------------------------------------

        for column_index, column in enumerate(
            columns,
            start=1
        ):

            letter = get_column_letter(
                column_index
            )

            max_length = len(
                str(column)
            )

            # Look at at most the first 1000
            # data rows for width calculation.

            for row in worksheet.iter_rows(
                min_row=2,
                max_row=min(
                    worksheet.max_row,
                    1001
                ),
                min_col=column_index,
                max_col=column_index
            ):

                value = row[0].value

                if value is None:
                    continue

                max_length = max(
                    max_length,
                    len(
                        str(value)
                    )
                )

            worksheet.column_dimensions[
                letter
            ].width = min(
                max(
                    max_length + 2,
                    12
                ),
                45
            )

        # -----------------------------------------------------
        # ROW HEIGHT
        # -----------------------------------------------------

        worksheet.row_dimensions[
            1
        ].height = 24

        # -----------------------------------------------------
        # WRAP TEXT
        # -----------------------------------------------------

        for row in worksheet.iter_rows():

            for cell in row:

                cell.alignment = Alignment(
                    vertical="top",
                    wrap_text=True
                )